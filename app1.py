"""
Excel Template Exporter with GitHub Sync

This Streamlit application provides:
1. Excel template creation and management
2. Master file processing with stock allocation
3. Advanced filtering with AND/OR logic
4. GitHub integration for data synchronization
5. Service restart capability

GITHUB SYNC FEATURE:
- Located in the upper left corner after login
- Three buttons available:
  * 📥 Pull from GitHub: Fetch latest changes from your repository
  * 🔄 Restart Services: Reload configurations and restart the app
  * 🚀 Pull & Restart: Combined action (recommended for updates)

SETUP:
1. Initialize git in your app directory: `git init`
2. Add remote repository: `git remote add origin <your-repo-url>`
3. Configure GITHUB_REPO_PATH in this file if needed
4. Use the sync buttons to pull updates and restart

WORKFLOW:
- Upload data files to GitHub
- Click "Pull & Restart" to sync and reload
- All templates and configurations will be updated automatically
"""

import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import io
import json
import os
from datetime import datetime
import subprocess
import time

# ============================================================================
# PAGE CONFIG
# ============================================================================
st.set_page_config(page_title="Excel Template Exporter", layout="wide")

# ============================================================================
# GITHUB CONFIGURATION
# ============================================================================
# Configure your GitHub repository path here
# Use "." for current directory, or specify full path to your repo
GITHUB_REPO_PATH = "."

# ============================================================================
# AUTHENTICATION
# ============================================================================
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "PaSSw0rd@1"

def check_login():
    """Check if user is logged in"""
    return st.session_state.get('logged_in', False)

def login_page():
    """Display login form"""
    st.title("🔐 Login Required")
    st.markdown("Please login to access the Excel Template Exporter")
    
    with st.form("login_form"):
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        submit = st.form_submit_button("Login", type="primary")
        
        if submit:
            if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
                st.session_state['logged_in'] = True
                st.success("Login successful!")
                st.rerun()
            else:
                st.error("Invalid username or password")

# Check authentication
if not check_login():
    login_page()
    st.stop()

# ============================================================================
# GITHUB PULL & SERVICE RESTART
# ============================================================================
def pull_from_github(repo_path=None):
    """
    Pull latest changes from GitHub repository
    
    Args:
        repo_path: Path to the git repository (default: uses GITHUB_REPO_PATH)
    
    Returns:
        tuple: (success: bool, message: str)
    """
    if repo_path is None:
        repo_path = GITHUB_REPO_PATH
    
    try:
        # Check if git is available
        result = subprocess.run(['git', '--version'], 
                              capture_output=True, 
                              text=True, 
                              timeout=5)
        
        if result.returncode != 0:
            return False, "⚠️ Git is not installed or not available"
        
        # Change to repo directory
        original_dir = os.getcwd()
        if repo_path != ".":
            if not os.path.exists(repo_path):
                return False, f"⚠️ Repository path does not exist: {repo_path}"
            os.chdir(repo_path)
        
        # Check if it's a git repository
        check_result = subprocess.run(['git', 'rev-parse', '--git-dir'], 
                                     capture_output=True, 
                                     text=True, 
                                     timeout=5)
        
        if check_result.returncode != 0:
            os.chdir(original_dir)
            return False, "⚠️ Not a git repository. Initialize git first with: git init"
        
        # Fetch latest changes
        fetch_result = subprocess.run(['git', 'fetch', 'origin'], 
                                     capture_output=True, 
                                     text=True, 
                                     timeout=30)
        
        if fetch_result.returncode != 0:
            os.chdir(original_dir)
            # Check if remote is configured
            if "fatal: 'origin' does not appear to be a git repository" in fetch_result.stderr:
                return False, "⚠️ No remote repository configured. Add remote with: git remote add origin <URL>"
            return False, f"⚠️ Git fetch failed: {fetch_result.stderr}"
        
        # Pull changes
        pull_result = subprocess.run(['git', 'pull', 'origin'], 
                                    capture_output=True, 
                                    text=True, 
                                    timeout=30)
        
        os.chdir(original_dir)
        
        if pull_result.returncode != 0:
            return False, f"⚠️ Git pull failed: {pull_result.stderr}"
        
        # Check if there were any changes
        if "Already up to date" in pull_result.stdout or "Already up-to-date" in pull_result.stdout:
            return True, "✅ Repository is already up to date"
        
        return True, f"✅ Successfully pulled from GitHub:\n{pull_result.stdout}"
        
    except subprocess.TimeoutExpired:
        return False, "⚠️ Git operation timed out"
    except Exception as e:
        return False, f"⚠️ Error during git pull: {str(e)}"

def restart_services():
    """
    Restart required services (Streamlit app reload)
    
    This will:
    1. Reload templates from disk
    2. Clear session state cache
    3. Trigger app rerun
    
    Returns:
        tuple: (success: bool, message: str)
    """
    try:
        # Reload templates from disk if templates.json was updated
        if os.path.exists(TEMPLATES_FILE):
            st.session_state['templates'] = load_templates()
        
        # Clear any cached data (optional - uncomment if needed)
        # st.cache_data.clear()
        # st.cache_resource.clear()
        
        return True, "✅ Services will restart now..."
    except Exception as e:
        return False, f"⚠️ Error restarting services: {str(e)}"

def render_github_controls():
    """Render GitHub pull and restart controls in the upper left corner"""
    st.markdown("### 🔄 GitHub Sync & Restart")
    
    # Show git status if available
    try:
        git_status = subprocess.run(['git', 'status', '--short'], 
                                   capture_output=True, 
                                   text=True, 
                                   timeout=5,
                                   cwd=GITHUB_REPO_PATH if GITHUB_REPO_PATH != "." else None)
        
        if git_status.returncode == 0:
            if git_status.stdout.strip():
                st.warning(f"⚠️ Uncommitted changes detected:\n```\n{git_status.stdout.strip()}\n```")
            else:
                # Get last commit info
                last_commit = subprocess.run(['git', 'log', '-1', '--pretty=format:%h - %s (%cr)'], 
                                           capture_output=True, 
                                           text=True, 
                                           timeout=5,
                                           cwd=GITHUB_REPO_PATH if GITHUB_REPO_PATH != "." else None)
                if last_commit.returncode == 0:
                    st.info(f"📝 Latest commit: {last_commit.stdout}")
    except:
        pass  # Silently ignore git status errors
    
    col1, col2, col3 = st.columns([2, 2, 3])
    
    with col1:
        if st.button("📥 Pull from GitHub", help="Pull latest changes from GitHub repository", use_container_width=True):
            with st.spinner("Pulling from GitHub..."):
                success, message = pull_from_github()
                
                if success:
                    st.success(message)
                    time.sleep(1.5)
                else:
                    st.error(message)
    
    with col2:
        if st.button("🔄 Restart Services", help="Restart all required services and reload configurations", use_container_width=True):
            with st.spinner("Restarting services..."):
                success, message = restart_services()
                
                if success:
                    st.success(message)
                    time.sleep(1)
                    st.rerun()
                else:
                    st.error(message)
    
    with col3:
        if st.button("🚀 Pull & Restart", type="primary", help="Pull from GitHub and restart services in one click", use_container_width=True):
            with st.spinner("Syncing and restarting..."):
                # First pull from GitHub
                pull_success, pull_msg = pull_from_github()
                
                if pull_success:
                    st.success(pull_msg)
                    time.sleep(1)
                    
                    # Then restart services
                    restart_success, restart_msg = restart_services()
                    if restart_success:
                        st.success(restart_msg)
                        time.sleep(1)
                        st.rerun()
                    else:
                        st.error(restart_msg)
                else:
                    st.error(pull_msg)
    
    st.divider()

# ============================================================================
# MASTER FILE PROCESSING
# ============================================================================
def convert_to_master_file(df_raw, filters=None):
    """
    Convert raw file to Master file with stock allocation logic.
    
    For items with same Item Name:
    - Sort by order date (SO Date)
    - Calculate running stock allocation
    - Update Stock Quantity to show available stock for each order
    
    Args:
        df_raw: Raw DataFrame to process
        filters: List of filter rules to apply before processing
                Each filter: {'column': str, 'contains': str, 'enabled': bool}
    """
    df = df_raw.copy()
    
    # ========================================================================
    # FILTER OUT UNWANTED ROWS BEFORE PROCESSING
    # ========================================================================
    if filters:
        initial_count = len(df)
        total_removed = 0
        
        # Apply each enabled filter
        for filter_rule in filters:
            if not filter_rule.get('enabled', True):
                continue
            
            column = filter_rule.get('column')
            search_text = filter_rule.get('contains', '').strip()
            
            if not search_text or column not in df.columns:
                continue
            
            # Apply filter (case-insensitive)
            before_count = len(df)
            mask = df[column].astype(str).str.lower().str.contains(search_text.lower(), na=False)
            df = df[~mask]
            removed = before_count - len(df)
            
            if removed > 0:
                st.info(f"🔍 Filtered out {removed} row(s) where '{column}' contains '{search_text}'")
                total_removed += removed
        
        if total_removed > 0:
            final_count = len(df)
            st.success(f"✅ Total rows removed: {total_removed} | Remaining: {final_count}")
    
    # ========================================================================
    
    # Ensure date column is datetime
    date_col = None
    for col in ['SO Date', 'Order Date', 'Date']:
        if col in df.columns:
            date_col = col
            break
    
    if date_col:
        df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
    
    # Determine grouping column (Item Code is more reliable than Item Name)
    group_col = None
    if 'Item Code' in df.columns:
        group_col = 'Item Code'
    elif 'Item Name' in df.columns:
        group_col = 'Item Name'
    else:
        st.warning("Neither 'Item Code' nor 'Item Name' found. Skipping Master File conversion.")
        return df
    
    # Sort by grouping column and Date
    sort_cols = [group_col]
    if date_col:
        sort_cols.append(date_col)
    df = df.sort_values(sort_cols, ascending=[True, True])
    
    # Group by Item Code/Name and calculate running stock
    result_rows = []
    
    for item_key, group in df.groupby(group_col, sort=False):
        # Get the original stock quantity (from first row of each item)
        original_stock = 0
        if 'Stock Quantity' in group.columns:
            # Find the first non-null, non-zero stock value
            stock_values = group['Stock Quantity'].dropna()
            if len(stock_values) > 0:
                original_stock = float(stock_values.iloc[0])
        
        remaining_stock = original_stock
        
        for idx, row in group.iterrows():
            # Use Pending Qty for allocation (not Order Qty!)
            # Pending Qty = what still needs to be fulfilled
            order_qty = 0
            if 'Pending Qty' in row.index:
                order_qty = float(row['Pending Qty']) if pd.notna(row['Pending Qty']) else 0
            elif 'Order Qty' in row.index:
                # Fallback to Order Qty if Pending Qty doesn't exist
                order_qty = float(row['Order Qty']) if pd.notna(row['Order Qty']) else 0
            
            # Calculate available stock for this order
            available_for_order = min(remaining_stock, order_qty) if remaining_stock > 0 else 0
            
            # Update the row
            row_copy = row.copy()
            
            # Update Stock Quantity to show allocated stock
            if 'Stock Quantity' in row_copy.index:
                row_copy['Stock Quantity'] = available_for_order
            
            # Add tracking columns (optional - can be hidden in export)
            row_copy['Original Stock'] = original_stock
            row_copy['Allocated Stock'] = available_for_order
            row_copy['Stock After Order'] = max(0, remaining_stock - order_qty)
            
            result_rows.append(row_copy)
            
            # Deduct from remaining stock
            remaining_stock = max(0, remaining_stock - order_qty)
    
    df_master = pd.DataFrame(result_rows)
    
    # Reset index
    df_master = df_master.reset_index(drop=True)
    
    # Add PRODUCTION column (Pending Qty - Stock Quantity) to master file
    if 'Pending Qty' in df_master.columns and 'Stock Quantity' in df_master.columns:
        df_master['PRODUCTION'] = df_master['Pending Qty'] - df_master['Stock Quantity']
    else:
        df_master['PRODUCTION'] = 0
    
    return df_master

# ============================================================================
# TEMPLATE MANAGEMENT
# ============================================================================
TEMPLATES_FILE = "templates.json"

def load_templates():
    if os.path.exists(TEMPLATES_FILE):
        try:
            with open(TEMPLATES_FILE, 'r') as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def save_templates(templates):
    with open(TEMPLATES_FILE, 'w') as f:
        json.dump(templates, f, indent=4)

if 'templates' not in st.session_state:
    st.session_state['templates'] = load_templates()

if 'filter_groups' not in st.session_state:
    st.session_state['filter_groups'] = []

# ============================================================================
# STYLING CONFIGURATION
# ============================================================================
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# ============================================================================
# FILTER & EXPORT LOGIC - ENHANCED WITH AND/OR SUPPORT
# ============================================================================
def apply_single_filter(df, col, op, val):
    """Apply a single filter condition to dataframe"""
    # Handle PRODUCTION column calculation if needed
    if col == 'PRODUCTION':
        if 'Pending Qty' in df.columns and 'Stock Quantity' in df.columns:
            df = df.copy()
            df['PRODUCTION'] = df['Pending Qty'] - df['Stock Quantity']
        else:
            return pd.Series([True] * len(df), index=df.index)
    
    if col not in df.columns:
        return df
    
    try:
        num_val = float(val)
        is_num = True
    except ValueError:
        num_val = val
        is_num = False
    
    if op == "==":
        if is_num:
            mask = df[col] == num_val
        else:
            search_val = str(val).strip().lower()
            mask = df[col].astype(str).str.strip().str.lower() == search_val
    elif op == "!=":
        if is_num:
            mask = df[col] != num_val
        else:
            search_val = str(val).strip().lower()
            mask = df[col].astype(str).str.strip().str.lower() != search_val
    elif op == ">" and is_num:
        mask = pd.to_numeric(df[col], errors='coerce') > num_val
    elif op == "<" and is_num:
        mask = pd.to_numeric(df[col], errors='coerce') < num_val
    elif op == ">=" and is_num:
        mask = pd.to_numeric(df[col], errors='coerce') >= num_val
    elif op == "<=" and is_num:
        mask = pd.to_numeric(df[col], errors='coerce') <= num_val
    elif op == "contains":
        mask = df[col].astype(str).str.contains(str(val), case=False, na=False)
    else:
        mask = pd.Series([True] * len(df), index=df.index)
    
    return mask

def get_filtered_dataframe(df_raw, config):
    """Apply complex filters with AND/OR logic support"""
    df_filtered = df_raw.copy()
    
    # Handle new format (filter_groups with AND/OR logic)
    filter_groups = config.get("filter_groups", [])
    
    # Backward compatibility: convert old "filters" format to new format
    if not filter_groups and "filters" in config and config["filters"]:
        filter_groups = [{
            "conditions": config["filters"],
            "logic": "AND"
        }]
    
    if filter_groups:
        # Apply filter groups
        group_masks = []
        
        for group in filter_groups:
            conditions = group.get("conditions", [])
            group_logic = group.get("logic", "AND")
            
            if not conditions:
                continue
            
            # Apply all conditions in this group
            condition_masks = []
            for f in conditions:
                if isinstance(f, dict) and "column" in f:
                    mask = apply_single_filter(df_raw, f["column"], f["operator"], f["value"])
                    condition_masks.append(mask)
            
            if not condition_masks:
                continue
            
            # Combine conditions within the group using group logic
            if group_logic == "OR":
                group_mask = condition_masks[0]
                for mask in condition_masks[1:]:
                    group_mask = group_mask | mask
            else:  # AND
                group_mask = condition_masks[0]
                for mask in condition_masks[1:]:
                    group_mask = group_mask & mask
            
            group_masks.append(group_mask)
        
        # Combine all group masks with the groups_logic (default AND)
        groups_logic = config.get("groups_logic", "AND")
        if group_masks:
            if groups_logic == "OR":
                final_mask = group_masks[0]
                for mask in group_masks[1:]:
                    final_mask = final_mask | mask
            else:  # AND
                final_mask = group_masks[0]
                for mask in group_masks[1:]:
                    final_mask = final_mask & mask
            
            # Apply INCLUSION filter - show only matching rows
            df_filtered = df_filtered[final_mask]
    
    # Add PRODUCTION column (Pending Qty - Stock Quantity) if requested
    if 'PRODUCTION' in config.get("columns", []):
        if 'Pending Qty' in df_filtered.columns and 'Stock Quantity' in df_filtered.columns:
            df_filtered['PRODUCTION'] = df_filtered['Pending Qty'] - df_filtered['Stock Quantity']
        else:
            df_filtered['PRODUCTION'] = 0
    
    # 2. Select columns (exclude tracking columns if not requested)
    available_cols = [c for c in config.get("columns", []) if c in df_filtered.columns]
    if available_cols:
        df_filtered = df_filtered[available_cols]
    
    # 3. Sort
    sort_cols = [c for c in config.get("sort_by", []) if c in df_filtered.columns]
    if sort_cols:
        df_filtered = df_filtered.sort_values(sort_cols)
        
    return df_filtered

def generate_excel_bytes(df_raw, specific_template_name=None):
    wb = Workbook()
    wb.remove(wb.active)
    
    templates_to_process = st.session_state['templates']
    if specific_template_name and specific_template_name in st.session_state['templates']:
        templates_to_process = {specific_template_name: st.session_state['templates'][specific_template_name]}
    
    if not templates_to_process:
        return None

    for template_name, config in templates_to_process.items():
        df_filtered = get_filtered_dataframe(df_raw, config)
        ws = wb.create_sheet(template_name[:31])
        
        columns = df_filtered.columns.tolist()
        for col_num, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER
        
        for r_idx, row in enumerate(dataframe_to_rows(df_filtered, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = BORDER
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                if isinstance(value, (int, float)) and not pd.isna(value):
                    if '.' in str(value):
                        cell.number_format = '0.00'
                    else:
                        cell.number_format = '0'
        
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 30)
            
        metadata_row = len(df_filtered) + 3
        ws[f'A{metadata_row}'] = f"Filter: {config.get('description', '')}"
        ws[f'A{metadata_row}'].font = Font(italic=True, color="666666", size=9)
        ws[f'A{metadata_row+1}'] = f"Records: {len(df_filtered)}"
        ws[f'A{metadata_row+1}'].font = Font(italic=True, color="666666", size=9)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def format_filter_display(filter_groups, groups_logic="AND"):
    """Create a human-readable display of filter logic"""
    if not filter_groups:
        return "No filters applied"
    
    parts = []
    for idx, group in enumerate(filter_groups):
        conditions = group.get("conditions", [])
        group_logic = group.get("logic", "AND")
        
        if not conditions:
            continue
        
        # Format conditions within group
        cond_strs = []
        for c in conditions:
            cond_strs.append(f"`{c['column']}` {c['operator']} `{c['value']}`")
        
        # Join conditions with group logic
        group_str = f" {group_logic} ".join(cond_strs)
        
        # Wrap in parentheses if more than one condition
        if len(cond_strs) > 1:
            group_str = f"({group_str})"
        
        parts.append(group_str)
    
    # Join groups with groups_logic
    if len(parts) > 1:
        return f" **{groups_logic}** ".join(parts)
    else:
        return parts[0] if parts else "No filters applied"

# ============================================================================
# UI LAYOUT
# ============================================================================

st.title("📊 Multi-Template Excel Exporter with Master File Processing")
st.markdown("Upload your raw Excel file → **Auto-convert to Master File with stock allocation** → Design custom templates with **AND/OR filter logic** → Export perfectly formatted multi-sheet Excel files.")

# Logout button
col_logout1, col_logout2 = st.columns([6, 1])
with col_logout2:
    if st.button("🚪 Logout"):
        st.session_state['logged_in'] = False
        st.rerun()

# GitHub Pull & Restart Controls
render_github_controls()

# 1. File Upload
uploaded_file = st.file_uploader("Upload Raw Excel File", type=["xlsx", "xls"])

if uploaded_file is not None:
    try:
        df_raw_original = pd.read_excel(uploaded_file)
        
        # ========================================================================
        # MASTER FILE FILTER CONFIGURATION
        # ========================================================================
        st.divider()
        st.subheader("🔍 Pre-Processing Filters (Optional)")
        st.markdown("Configure filters to remove unwanted rows **before** creating the master file:")
        
        with st.expander("⚙️ Configure Filters", expanded=False):
            # Initialize session state for master filters
            if 'master_filters' not in st.session_state:
                st.session_state['master_filters'] = [
                    {'column': "Party's Order No.", 'contains': 'from stock', 'enabled': True},
                    {'column': 'Line', 'contains': 'trading', 'enabled': True}
                ]
            
            # Display current filters
            st.markdown("**Filter Rules:**")
            
            filters_to_remove = []
            for i, filter_rule in enumerate(st.session_state['master_filters']):
                col1, col2, col3, col4 = st.columns([1, 3, 3, 1])
                
                with col1:
                    filter_rule['enabled'] = st.checkbox("✓", value=filter_rule['enabled'], key=f"filter_enable_{i}")
                
                with col2:
                    available_cols = list(df_raw_original.columns)
                    current_col = filter_rule['column'] if filter_rule['column'] in available_cols else available_cols[0]
                    filter_rule['column'] = st.selectbox("Column", options=available_cols, index=available_cols.index(current_col), key=f"filter_col_{i}", label_visibility="collapsed")
                
                with col3:
                    filter_rule['contains'] = st.text_input("Contains (case-insensitive)", value=filter_rule['contains'], key=f"filter_text_{i}", placeholder="e.g., from stock", label_visibility="collapsed")
                
                with col4:
                    if st.button("🗑️", key=f"filter_del_{i}"):
                        filters_to_remove.append(i)
            
            # Remove filters marked for deletion
            for idx in reversed(filters_to_remove):
                st.session_state['master_filters'].pop(idx)
                st.rerun()
            
            # Action buttons
            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                if st.button("➕ Add Filter Rule"):
                    st.session_state['master_filters'].append({
                        'column': list(df_raw_original.columns)[0],
                        'contains': '',
                        'enabled': True
                    })
                    st.rerun()
            
            with col_btn2:
                if st.button("🔄 Reset to Defaults"):
                    st.session_state['master_filters'] = [
                        {'column': "Party's Order No.", 'contains': 'from stock', 'enabled': True},
                        {'column': 'Line', 'contains': 'trading', 'enabled': True}
                    ]
                    st.rerun()
        
        # Show summary of active filters
        active_filters = [f for f in st.session_state['master_filters'] if f['enabled'] and f['contains'].strip()]
        if active_filters:
            st.info(f"✅ {len(active_filters)} filter rule(s) will be applied")
            for f in active_filters:
                st.markdown(f"- Remove rows where **{f['column']}** contains *'{f['contains']}'*")
        else:
            st.warning("⚠️ No filters active - all rows will be processed")
        
        st.divider()
        # ========================================================================
        
        # CONVERT TO MASTER FILE
        with st.spinner("🔄 Converting to Master File (calculating stock allocation)..."):
            df_raw = convert_to_master_file(df_raw_original, st.session_state.get('master_filters', []))
        
        st.success(f"✅ Master File ready: {len(df_raw)} rows and {len(df_raw.columns)} columns.")
        
        # Show conversion summary
        with st.expander("📋 Master File Conversion Summary", expanded=False):
            col_sum1, col_sum2, col_sum3 = st.columns(3)
            with col_sum1:
                st.metric("Total Rows", len(df_raw))
            with col_sum2:
                # Show unique items based on grouping column used
                group_col = 'Item Code' if 'Item Code' in df_raw.columns else 'Item Name'
                if group_col in df_raw.columns:
                    st.metric(f"Unique {group_col}s", df_raw[group_col].nunique())
            with col_sum3:
                if 'Stock Quantity' in df_raw.columns:
                    st.metric("Total Allocated Stock", f"{df_raw['Stock Quantity'].sum():.0f}")
            
            st.markdown("**Master File Logic Applied:**")
            st.markdown(f"""
            - Items grouped by **{group_col}** (unique identifier)
            - Sorted by **SO Date** (earliest first within each item)
            - **Stock Quantity** calculated using **Pending Qty** (not Order Qty)
            - Each order shows **available stock** for pending quantity after previous orders
            - Logic: Allocates stock chronologically based on what still needs to be fulfilled
            """)
        
        with st.expander("📊 Master File Preview", expanded=False):
            # Show relevant columns for preview
            preview_cols = [c for c in df_raw.columns if c not in ['Original Stock', 'Stock After Order']]
            st.dataframe(df_raw[preview_cols].head(100), use_container_width=True)
        
        # Download Master File button
        st.markdown("### 📥 Download Master File")
        col_download1, col_download2 = st.columns([2, 1])
        with col_download1:
            st.markdown("Download the processed Master File with stock allocation calculations.")
        with col_download2:
            # Generate Excel for Master File
            master_wb = Workbook()
            master_ws = master_wb.active
            master_ws.title = "Master File"
            
            # Add headers
            for col_num, col_name in enumerate(df_raw.columns, 1):
                cell = master_ws.cell(row=1, column=col_num, value=col_name)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = BORDER
            
            # Add data
            for r_idx, row in enumerate(dataframe_to_rows(df_raw, index=False, header=False), 2):
                for c_idx, value in enumerate(row, 1):
                    cell = master_ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = BORDER
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
                    if isinstance(value, (int, float)) and not pd.isna(value):
                        if '.' in str(value):
                            cell.number_format = '0.00'
                        else:
                            cell.number_format = '0'
            
            # Auto-adjust column widths
            for col in master_ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                master_ws.column_dimensions[col_letter].width = min(max_length + 2, 30)
            
            # Add metadata
            metadata_row = len(df_raw) + 3
            master_ws[f'A{metadata_row}'] = f"Master File Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            master_ws[f'A{metadata_row}'].font = Font(italic=True, color="666666", size=9)
            master_ws[f'A{metadata_row+1}'] = f"Total Records: {len(df_raw)}"
            master_ws[f'A{metadata_row+1}'].font = Font(italic=True, color="666666", size=9)
            
            # Save to bytes
            master_output = io.BytesIO()
            master_wb.save(master_output)
            master_output.seek(0)
            
            st.download_button(
                label="⬇️ Download Master File",
                data=master_output,
                file_name=f"Master_File_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )
            
    except Exception as e:
        st.error(f"Error reading file: {e}")
        st.stop()
        
    st.divider()
    
    # Split layout into Designer and Saved Templates
    col1, col2 = st.columns([1, 1])
    
    # --- DESIGNER (Left Column) ---
    with col1:
        st.subheader("🛠️ Template Designer")
        
        t_name = st.text_input("Template Name")
        
        # All columns are available (PRODUCTION is added to master file)
        selectable_cols = list(df_raw.columns)
        
        # Exclude internal tracking columns from default selection
        default_cols = [c for c in selectable_cols if c not in ['Original Stock', 'Allocated Stock', 'Stock After Order']]
        
        t_cols = st.multiselect("Select Columns to Include", options=selectable_cols, default=default_cols)
        
        # Info about available columns
        st.info("ℹ️ **PRODUCTION Column**: Automatically calculated as Pending Qty - Stock Quantity (Available for all templates)")
        t_sort = st.selectbox("Sort By (Optional)", options=["None"] + list(df_raw.columns))
        
        st.markdown("---")
        st.markdown("### 🔍 Advanced Filters (AND/OR Logic)")
        
        # Current filter groups display
        if st.session_state['filter_groups']:
            st.markdown("**Current Filter Logic:**")
            
            # Only show AND/OR option if there are 2+ filter groups
            groups_logic = "AND"  # Default
            if len(st.session_state['filter_groups']) > 1:
                groups_logic = st.radio(
                    "Combine filter groups with:",
                    options=["AND", "OR"],
                    horizontal=True,
                    key="groups_logic_radio"
                )
            
            filter_display = format_filter_display(st.session_state['filter_groups'], groups_logic)
            st.markdown(filter_display, unsafe_allow_html=True)
            
            if st.button("🗑️ Clear All Filters"):
                st.session_state['filter_groups'] = []
                st.rerun()
        
        st.markdown("---")
        st.markdown("**Create Filter Group:**")
        
        # Initialize current group in session state
        if 'current_group_conditions' not in st.session_state:
            st.session_state['current_group_conditions'] = []
        
        # Add condition to current group
        f_col1, f_col2, f_col3 = st.columns([2, 1, 2])
        filter_col = f_col1.selectbox("Column", options=df_raw.columns, key="f_col")
        filter_op = f_col2.selectbox("Operator", options=["==", "!=", ">", "<", ">=", "<=", "contains"], key="f_op")
        filter_val = f_col3.text_input("Value", key="f_val")
        
        col_add1, col_add2 = st.columns(2)
        with col_add1:
            if st.button("➕ Add to Group"):
                if filter_val != "":
                    st.session_state['current_group_conditions'].append({
                        "column": filter_col,
                        "operator": filter_op,
                        "value": filter_val
                    })
                    st.rerun()
                else:
                    st.warning("Please enter a value for the filter.")
        
        # Display current group being built
        if st.session_state['current_group_conditions']:
            st.markdown("**Current Group Conditions:**")
            for i, c in enumerate(st.session_state['current_group_conditions']):
                col_disp1, col_disp2 = st.columns([4, 1])
                with col_disp1:
                    st.write(f"{i+1}. `{c['column']}` {c['operator']} `{c['value']}`")
                with col_disp2:
                    if st.button("❌", key=f"del_cond_{i}"):
                        st.session_state['current_group_conditions'].pop(i)
                        st.rerun()
            
            # Only show AND/OR option if there are 2+ conditions
            group_logic = "AND"  # Default
            if len(st.session_state['current_group_conditions']) > 1:
                group_logic = st.radio(
                    "Combine conditions in this group with:",
                    options=["AND", "OR"],
                    horizontal=True,
                    key="group_logic_radio"
                )
            
            col_save1, col_save2 = st.columns(2)
            with col_save1:
                if st.button("✅ Save Filter Group", type="primary"):
                    st.session_state['filter_groups'].append({
                        "conditions": st.session_state['current_group_conditions'].copy(),
                        "logic": group_logic
                    })
                    st.session_state['current_group_conditions'] = []
                    st.rerun()
            with col_save2:
                if st.button("🔄 Reset Group"):
                    st.session_state['current_group_conditions'] = []
                    st.rerun()
        
        st.markdown("---")
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("💾 Save Complete Template", type="primary"):
            if not t_name:
                st.error("Template Name is required!")
            elif not t_cols:
                st.error("At least one column must be selected!")
            else:
                config = {
                    "description": f"Custom template: {t_name}",
                    "columns": t_cols,
                    "filter_groups": st.session_state['filter_groups'].copy(),
                    "groups_logic": st.session_state.get('groups_logic_radio', 'AND'),
                    "sort_by": [t_sort] if t_sort != "None" else []
                }
                st.session_state['templates'][t_name] = config
                save_templates(st.session_state['templates'])
                
                # Clear form state
                st.session_state['filter_groups'] = []
                st.session_state['current_group_conditions'] = []
                st.success(f"Template '{t_name}' saved successfully!")
                st.rerun()

    # --- SAVED TEMPLATES (Right Column) ---
    with col2:
        st.subheader("📁 Saved Templates")
        
        if not st.session_state['templates']:
            st.info("No templates saved yet. Create one on the left!")
        else:
            selected_t = st.selectbox("Select a Template to Preview", options=list(st.session_state['templates'].keys()))
            
            if selected_t:
                config = st.session_state['templates'][selected_t]
                
                # Display filter logic
                if config.get("filter_groups"):
                    st.markdown("**Filter Logic:**")
                    filter_display = format_filter_display(
                        config.get("filter_groups", []),
                        config.get("groups_logic", "AND")
                    )
                    st.markdown(filter_display, unsafe_allow_html=True)
                    st.markdown("---")
                
                filtered_df = get_filtered_dataframe(df_raw, config)
                
                st.write(f"**Rows Matched:** {len(filtered_df)}")
                
                if len(filtered_df) == 0 and config.get("filter_groups"):
                    st.warning("Your filter matched 0 rows. Here is some sample data from the columns you filtered to help debug:")
                    for group in config.get("filter_groups", []):
                        for f in group.get("conditions", []):
                            c = f.get("column")
                            if c in df_raw.columns:
                                uniq = df_raw[c].dropna().unique()
                                st.write(f"- `{c}`: {', '.join([str(x) for x in uniq[:10]])}")
                
                st.dataframe(filtered_df.head(50), use_container_width=True)
                
                if st.button("🗑️ Delete Template"):
                    del st.session_state['templates'][selected_t]
                    save_templates(st.session_state['templates'])
                    st.rerun()
                
                st.divider()
                st.subheader("📤 Export")
                
                # EXPORT SINGLE
                if len(filtered_df) > 0:
                    excel_data_single = generate_excel_bytes(df_raw, specific_template_name=selected_t)
                    st.download_button(
                        label=f"⬇️ Export SELECTED ({selected_t})",
                        data=excel_data_single,
                        file_name=f"{selected_t}_export.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
                # EXPORT ALL
                st.markdown("<br>", unsafe_allow_html=True)
                excel_data_all = generate_excel_bytes(df_raw)
                st.download_button(
                    label="⬇️ Export ALL Templates (Multi-sheet)",
                    data=excel_data_all,
                    file_name="All_Templates_Export.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )
