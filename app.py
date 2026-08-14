"""
Excel Master File & Template Exporter - FINAL COMPLETE v4.0

FEATURES:
✅ Create Master File from Source
✅ Generate 10 Sheets with exact filters
✅ Create COMBINED file (all sheets in one workbook)
✅ Create SEPARATE files (each sheet as individual .xlsx)
✅ Dynamic Template Management (Add/Edit/Delete sheets)
✅ Flexible configuration system
✅ Date sorting + Subtotals + TOTAL rows
✅ Professional Excel formatting
"""

import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import io
import json
import os
import copy
from datetime import datetime
import zipfile
import subprocess
import sys
import time
import hashlib

# ============================================================================
# PAGE CONFIG
# ============================================================================
st.set_page_config(page_title="Excel Master & Template Exporter v4.8", layout="wide")

# ============================================================================
# AUTO-VERSION MANAGEMENT (NEW FEATURE)
# ============================================================================
VERSION_FILE = "version.json"
CHANGELOG_FILE = "changelog.json"

def get_file_hash(file_path):
    """Get SHA256 hash of a file"""
    try:
        if not os.path.exists(file_path):
            return None
        sha256_hash = hashlib.sha256()
        with open(file_path, "rb") as f:
            for byte_block in iter(lambda: f.read(4096), b""):
                sha256_hash.update(byte_block)
        return sha256_hash.hexdigest()
    except:
        return None

def load_version_info():
    """Load version info from file"""
    try:
        if os.path.exists(VERSION_FILE):
            with open(VERSION_FILE, 'r') as f:
                return json.load(f)
    except:
        pass

    return {
        "major": 4,
        "minor": 8,
        "patch": 0,
        "build": 1,
        "created": datetime.now().isoformat(),
        "last_modified": datetime.now().isoformat(),
        "app_hash": get_file_hash(__file__),
        "config_hash": None,
        "templates_hash": None
    }

def save_version_info(version_info):
    """Save version info"""
    try:
        with open(VERSION_FILE, 'w') as f:
            json.dump(version_info, f, indent=2)
        return True
    except:
        return False

def load_changelog():
    """Load changelog"""
    try:
        if os.path.exists(CHANGELOG_FILE):
            with open(CHANGELOG_FILE, 'r') as f:
                return json.load(f)
    except:
        pass
    return []

def save_changelog(changelog):
    """Save changelog"""
    try:
        with open(CHANGELOG_FILE, 'w') as f:
            json.dump(changelog, f, indent=2)
        return True
    except:
        return False

def add_changelog_entry(change_type, description, details=""):
    """Add changelog entry"""
    try:
        changelog = load_changelog()
        entry = {
            "timestamp": datetime.now().isoformat(),
            "type": change_type,
            "description": description,
            "details": details,
            "version": get_version_string(),
            "user": "system"
        }
        changelog.insert(0, entry)
        changelog = changelog[:100]
        save_changelog(changelog)
    except:
        pass

def get_version_string():
    """Get version string"""
    try:
        if 'version_info' in st.session_state:
            v = st.session_state.version_info
            return f"v{v['major']}.{v['minor']}.{v['patch']}.{v['build']}"
    except:
        pass
    return "v4.8.0.1"

def check_for_changes():
    """Check if files changed"""
    try:
        if 'version_info' not in st.session_state:
            return False, []

        version_info = st.session_state.version_info
        changed = False
        change_list = []

        # Check app.py
        app_hash = get_file_hash(__file__)
        if app_hash and app_hash != version_info.get('app_hash'):
            changed = True
            change_list.append("App code modified")
            version_info['app_hash'] = app_hash

        # Check templates
        if os.path.exists(TEMPLATES_FILE):
            template_hash = get_file_hash(TEMPLATES_FILE)
            if template_hash and template_hash != version_info.get('templates_hash'):
                changed = True
                change_list.append("Templates modified")
                version_info['templates_hash'] = template_hash

        # Check config
        if os.path.exists(CUSTOM_TEMPLATES_FILE):
            config_hash = get_file_hash(CUSTOM_TEMPLATES_FILE)
            if config_hash and config_hash != version_info.get('config_hash'):
                changed = True
                change_list.append("Configuration modified")
                version_info['config_hash'] = config_hash

        return changed, change_list
    except:
        return False, []

def increment_version(version_info, change_list):
    """Increment version"""
    try:
        if change_list:
            if any('code' in c.lower() for c in change_list):
                version_info['minor'] += 1
                version_info['patch'] = 0
                version_info['build'] = 1
            else:
                version_info['patch'] += 1
                version_info['build'] = 1
        else:
            version_info['build'] += 1

        version_info['last_modified'] = datetime.now().isoformat()
        save_version_info(version_info)
    except:
        pass

# Initialize version info
if 'version_info' not in st.session_state:
    st.session_state.version_info = load_version_info()

# Check for changes
try:
    changed, change_list = check_for_changes()
    if changed:
        increment_version(st.session_state.version_info, change_list)
except:
    pass

# ============================================================================
# GITHUB CONFIG
# ============================================================================
GITHUB_CONFIG = {
    'repo_url': 'https://github.com/your-username/excel-exporter.git',
    'branch': 'main',
    'enabled': False  # Set to True to enable GitHub sync
}

# ============================================================================
# AUTHENTICATION
# ============================================================================
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "PaSSw0rd@1"
TEMPLATES_FILE = "sheet_templates.json"
CUSTOM_TEMPLATES_FILE = "custom_templates.json"

def check_login():
    return st.session_state.get('logged_in', False)

def login_page():
    st.title("🔐 Login Required")
    st.markdown("Excel Master & Template Exporter v4.9")

    with st.form("login_form"):
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        submit = st.form_submit_button("Login", type="primary")

        if submit:
            if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
                st.session_state['logged_in'] = True
                st.success("Login successful!")
                st.rerun()
            else:
                st.error("Invalid username or password")

if not check_login():
    login_page()
    st.stop()

# ============================================================================
# GITHUB SYNC & RESTART FUNCTIONS
# ============================================================================
def get_git_status():
    """Get current Git status"""
    try:
        result = subprocess.run(
            ['git', 'status', '--porcelain'],
            capture_output=True,
            text=True,
            timeout=10
        )
        return result.stdout.strip() if result.returncode == 0 else None
    except:
        return None

def get_git_log(limit=5):
    """Get recent Git commits"""
    try:
        result = subprocess.run(
            ['git', 'log', f'--oneline', f'-{limit}'],
            capture_output=True,
            text=True,
            timeout=10
        )
        return result.stdout.strip() if result.returncode == 0 else None
    except:
        return None

def get_current_branch():
    """Get current Git branch"""
    try:
        result = subprocess.run(
            ['git', 'rev-parse', '--abbrev-ref', 'HEAD'],
            capture_output=True,
            text=True,
            timeout=10
        )
        return result.stdout.strip() if result.returncode == 0 else "unknown"
    except:
        return "unknown"

def get_current_commit():
    """Get current Git commit hash (short)"""
    try:
        result = subprocess.run(
            ['git', 'rev-parse', '--short', 'HEAD'],
            capture_output=True,
            text=True,
            timeout=10
        )
        return result.stdout.strip() if result.returncode == 0 else "unknown"
    except:
        return "unknown"

def get_uncommitted_changes():
    """Get count of uncommitted changes"""
    try:
        result = subprocess.run(
            ['git', 'status', '--porcelain'],
            capture_output=True,
            text=True,
            timeout=10
        )
        return len(result.stdout.strip().split('\n')) if result.stdout.strip() else 0
    except:
        return 0

def sync_with_github():
    """Sync application with GitHub repository"""
    try:
        progress_bar = st.progress(0)
        status_text = st.empty()

        status_text.text("🔄 Fetching from GitHub...")
        progress_bar.progress(25)

        # Fetch updates
        fetch_result = subprocess.run(
            ['git', 'fetch', 'origin', GITHUB_CONFIG['branch']],
            capture_output=True,
            text=True,
            timeout=30
        )

        progress_bar.progress(50)
        status_text.text("📥 Pulling latest changes...")

        # Pull latest changes
        result = subprocess.run(
            ['git', 'pull', 'origin', GITHUB_CONFIG['branch']],
            capture_output=True,
            text=True,
            timeout=30
        )

        progress_bar.progress(100)

        if result.returncode == 0:
            st.success("✅ GitHub sync successful!")
            st.balloons()
            st.info(f"Latest changes from '{GITHUB_CONFIG['branch']}' branch pulled.")
            return True
        else:
            st.error(f"❌ Git pull failed: {result.stderr}")
            return False
    except subprocess.TimeoutExpired:
        st.error("❌ GitHub sync timed out (exceeded 30 seconds)")
        return False
    except Exception as e:
        st.error(f"❌ GitHub sync error: {str(e)}")
        return False

def push_to_github(message="Update from Excel Exporter"):
    """Push changes to GitHub repository"""
    try:
        progress_bar = st.progress(0)
        status_text = st.empty()

        status_text.text("📝 Staging changes...")
        progress_bar.progress(25)

        # Stage changes
        subprocess.run(['git', 'add', '.'], capture_output=True, timeout=10)

        progress_bar.progress(50)
        status_text.text("💾 Creating commit...")

        # Commit changes
        result = subprocess.run(
            ['git', 'commit', '-m', message],
            capture_output=True,
            text=True,
            timeout=10
        )

        progress_bar.progress(75)
        status_text.text("📤 Pushing to GitHub...")

        # Push changes
        push_result = subprocess.run(
            ['git', 'push', 'origin', GITHUB_CONFIG['branch']],
            capture_output=True,
            text=True,
            timeout=30
        )

        progress_bar.progress(100)

        if push_result.returncode == 0:
            st.success("✅ Successfully pushed to GitHub!")
            st.balloons()
            return True
        else:
            st.warning("⚠️ Push may have encountered issues")
            st.write(push_result.stderr)
            return False
    except Exception as e:
        st.error(f"❌ GitHub push error: {str(e)}")
        return False

def pull_and_restart():
    """Pull from GitHub and restart application"""
    try:
        progress_bar = st.progress(0)
        status_text = st.empty()

        # Step 1: Fetch
        status_text.text("🔄 Step 1: Fetching from GitHub...")
        progress_bar.progress(15)

        fetch_result = subprocess.run(
            ['git', 'fetch', 'origin', GITHUB_CONFIG['branch']],
            capture_output=True,
            text=True,
            timeout=30
        )

        # Step 2: Pull
        progress_bar.progress(40)
        status_text.text("📥 Step 2: Pulling latest changes...")

        pull_result = subprocess.run(
            ['git', 'pull', 'origin', GITHUB_CONFIG['branch']],
            capture_output=True,
            text=True,
            timeout=30
        )

        if pull_result.returncode != 0:
            st.error(f"❌ Git pull failed: {pull_result.stderr}")
            return

        progress_bar.progress(70)
        status_text.text("✅ Step 3: Latest code downloaded!")

        # Step 3: Restart
        progress_bar.progress(90)
        status_text.text("🔄 Step 4: Restarting application...")
        st.success("✅ GitHub pull successful!")
        st.info("🔄 Restarting app in 2 seconds...")
        st.balloons()

        progress_bar.progress(100)

        # Clear session and restart
        import time
        time.sleep(2)

        for key in list(st.session_state.keys()):
            del st.session_state[key]

        st.rerun()

    except subprocess.TimeoutExpired:
        st.error("❌ Operation timed out (exceeded 30 seconds)")
    except Exception as e:
        st.error(f"❌ Error: {str(e)}")

def restart_application():
    """Restart the Streamlit application"""
    progress_bar = st.progress(0)
    status_text = st.empty()

    status_text.text("🔄 Clearing session state...")
    progress_bar.progress(50)

    st.warning("🔄 Restarting application in 2 seconds...")
    st.balloons()

    progress_bar.progress(90)
    status_text.text("🔄 Restarting...")

    # Clear session state
    import time
    time.sleep(2)

    for key in list(st.session_state.keys()):
        del st.session_state[key]

    progress_bar.progress(100)
    # Rerun the app
    st.rerun()

def rollback_changes():
    """Rollback to previous Git commit"""
    try:
        result = subprocess.run(
            ['git', 'reset', '--hard', 'HEAD~1'],
            capture_output=True,
            text=True,
            timeout=10
        )

        if result.returncode == 0:
            st.success("✅ Rolled back to previous commit!")
            st.info("Previous changes have been undone.")
            return True
        else:
            st.error(f"❌ Rollback failed: {result.stderr}")
            return False
    except Exception as e:
        st.error(f"❌ Rollback error: {str(e)}")
        return False

# ============================================================================
# SYSTEM STATUS SIDEBAR
# ============================================================================
def show_system_status():
    """Display system status in sidebar"""
    with st.sidebar:
        st.divider()
        st.subheader("⚙️ System Controls")

        # Main combined button
        if GITHUB_CONFIG['enabled']:
            if st.button("🚀 Pull & Restart", use_container_width=True, key="pull_restart_btn"):
                pull_and_restart()
        else:
            if st.button("🔄 Restart", use_container_width=True, key="restart_btn"):
                restart_application()

        # Individual buttons (if GitHub enabled)
        if GITHUB_CONFIG['enabled']:
            st.divider()
            st.caption("**Other Options:**")

            col1, col2, col3 = st.columns(3)

            with col1:
                if st.button("📥 Sync", use_container_width=True, key="sync_btn"):
                    sync_with_github()

            with col2:
                if st.button("🔄 Restart Only", use_container_width=True, key="restart_only_btn"):
                    restart_application()

            with col3:
                if st.button("📊 Status", use_container_width=True, key="status_btn"):
                    st.session_state['show_git_status'] = True

        st.divider()

        # Display app version and status
        st.caption("📊 Excel Master & Template Exporter")
        st.caption("Version: 4.8")
        st.caption(f"Status: ✅ Running")
        st.caption(f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        if GITHUB_CONFIG['enabled']:
            current_branch = get_current_branch()
            uncommitted = get_uncommitted_changes()

            st.caption(f"🔗 GitHub: {GITHUB_CONFIG['repo_url'].split('/')[-1]}")
            st.caption(f"📌 Branch: {current_branch}")

            if uncommitted > 0:
                st.caption(f"⚠️ Changes: {uncommitted} file(s)")
            else:
                st.caption(f"✅ Clean workspace")

# ============================================================================
# GITHUB STATUS DASHBOARD
# ============================================================================
def show_github_dashboard():
    """Show comprehensive GitHub sync status dashboard"""
    st.header("📊 GitHub Sync Dashboard")

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        branch = get_current_branch()
        st.metric("📌 Current Branch", branch)

    with col2:
        changes = get_uncommitted_changes()
        st.metric("⚠️ Uncommitted Changes", changes)

    with col3:
        st.metric("🔗 Repository", GITHUB_CONFIG['repo_url'].split('/')[-1].replace('.git', ''))

    with col4:
        st.metric("⏱️ Last Check", datetime.now().strftime('%H:%M:%S'))

    st.divider()

    # Git status details
    st.subheader("📝 Git Status Details")
    git_status = get_git_status()
    if git_status:
        st.code(git_status, language="bash")
    else:
        st.info("✅ Working directory clean")

    st.divider()

    # Recent commits
    st.subheader("📜 Recent Commits")
    git_log = get_git_log(10)
    if git_log:
        st.code(git_log, language="bash")
    else:
        st.warning("No commits found")

    st.divider()

    # GitHub operations
    st.subheader("🔄 GitHub Operations")
    col1, col2, col3 = st.columns(3)

    with col1:
        if st.button("📥 Pull Latest", use_container_width=True):
            sync_with_github()

    with col2:
        commit_msg = st.text_input("Commit message", "Update from Excel Exporter")
        if st.button("📤 Push Changes", use_container_width=True):
            push_to_github(commit_msg)

    with col3:
        if st.button("↩️ Rollback", use_container_width=True):
            st.warning("⚠️ Are you sure? This will undo the last commit.")
            col_rb1, col_rb2 = st.columns(2)
            with col_rb1:
                if st.button("✅ Yes, Rollback", use_container_width=True):
                    rollback_changes()
            with col_rb2:
                if st.button("❌ Cancel", use_container_width=True):
                    st.info("Rollback cancelled")

# ============================================================================
# DEFAULT SHEET TEMPLATES
# ============================================================================
DEFAULT_SHEETS = {
    'SAMPLE': {
        'filter_type': 'remarks',
        'filter_value': 'SAMPLE',
        'columns': 13,
        'has_subtotals': False,
        'enabled': True,
        'description': 'Sample orders'
    },
    'HOLD': {
        'filter_type': 'remarks',
        'filter_value': 'Hold',
        'columns': 13,
        'has_subtotals': False,
        'enabled': True,
        'description': 'Hold orders (remarks contains Hold)'
    },
    'RU': {
        'filter_type': 'sales_person_single',
        'filter_value': 'RU',
        'exclude_remarks': ['Hold', 'SAMPLE'],  # Exclude Hold and Sample
        'columns': 19,
        'has_subtotals': False,
        'enabled': True,
        'description': 'RU Production Division (excluding Hold & Sample)'
    },
    'GS': {
        'filter_type': 'sales_person_single',
        'filter_value': 'GS',
        'exclude_remarks': ['Hold', 'SAMPLE'],  # Exclude Hold and Sample
        'columns': 19,
        'has_subtotals': False,
        'enabled': True,
        'description': 'GS Production Division (excluding Hold & Sample)'
    },
    'NKG': {
        'filter_type': 'sales_person_single',
        'filter_value': 'NKG',
        'exclude_remarks': ['Hold', 'SAMPLE'],  # Exclude Hold and Sample
        'columns': 19,
        'has_subtotals': False,
        'enabled': True,
        'description': 'NKG Production Division (excluding Hold & Sample)'
    },
    'AVS': {
        'filter_type': 'sales_person_grouped',
        'filter_value': ['ARUN', 'JS', 'PS'],
        'exclude_remarks': ['Hold', 'SAMPLE'],  # Exclude Hold and Sample
        'columns': 19,
        'has_subtotals': False,  # FIXED: Don't count subtotal rows
        'group_by': 'Sales Person Name',
        'enabled': True,
        'description': 'AVS with Sales Person subtotals (excluding Hold & Sample)'
    },
    'NITIN': {
        'filter_type': 'none',
        'exclude_remarks': ['Hold', 'SAMPLE'],
        'columns': 13,
        'has_subtotals': False,
        'enabled': True,
        'description': 'Semi Finished Good only (excluding Raw Material, Finished Good, Hold & Sample)',
        # ✅ PRE-PROCESSING FILTERS for Nitin - KEEP ONLY Semi Finished Good
        'pre_processing_filters': [
            {'column': 'Category', 'contains': '^Raw Material$', 'regex': True, 'enabled': True},
            {'column': 'Category', 'contains': '^Finished Good$', 'regex': True, 'enabled': True},
            {'column': 'Remarks', 'contains': 'Sample', 'enabled': True},
            {'column': 'Remarks', 'contains': 'Hold', 'enabled': True}
        ]
    },
    'RAJESH': {
        'filter_type': 'line_grouped',
        'filter_value': ['PHILIPS', 'PHILLIPS', 'OTG', 'IMM.ROD'],
        'exclude_remarks': ['Hold', 'SAMPLE'],
        'columns': 13,
        'has_subtotals': True,
        'group_by': 'Line',
        'enabled': True,
        'description': 'PHILIPS/PHILLIPS Air Fryer (excluding Semi Finished Good & Hold & Sample)',
        # ✅ PRE-PROCESSING FILTERS for Rajesh
        'pre_processing_filters': [
            {'column': 'Category', 'contains': 'Semi Finished Good', 'enabled': True},
            {'column': 'Remarks', 'contains': 'Sample', 'enabled': True},
            {'column': 'Remarks', 'contains': 'Hold', 'enabled': True}
        ]
    },
    'PRADEEP': {
        'filter_type': 'line_grouped',
        'filter_value': ['LINE C', 'MARKET'],
        'exclude_remarks': ['Hold', 'SAMPLE'],
        'columns': 13,
        'has_subtotals': True,
        'group_by': 'Line',
        'enabled': True,
        'description': 'LINE C + MARKET (excluding Semi Finished Good & Hold & Sample)',
        # ✅ PRE-PROCESSING FILTERS for Pradeep
        'pre_processing_filters': [
            {'column': 'Category', 'contains': 'Semi Finished Good', 'enabled': True},
            {'column': 'Remarks', 'contains': 'Sample', 'enabled': True},
            {'column': 'Remarks', 'contains': 'Hold', 'enabled': True}
        ]
    },
    'ASHISH': {
        'filter_type': 'line_grouped',
        'filter_value': ['LINE G/L', 'LINE SS'],
        'exclude_remarks': ['Hold', 'SAMPLE'],
        'columns': 13,
        'has_subtotals': True,
        'group_by': 'Line',
        'enabled': True,
        'description': 'LINE G/L + LINE SS (excluding Semi Finished Good & Hold & Sample)',
        # ✅ PRE-PROCESSING FILTERS for Ashish
        'pre_processing_filters': [
            {'column': 'Category', 'contains': 'Semi Finished Good', 'enabled': True},
            {'column': 'Remarks', 'contains': 'Sample', 'enabled': True},
            {'column': 'Remarks', 'contains': 'Hold', 'enabled': True}
        ]
    },
    'Master File': {
        'filter_type': 'none',
        'columns': 19,
        'has_subtotals': False,
        'enabled': True,
        'description': 'Complete master data'
    }
}

# ============================================================================
# TEMPLATE MANAGEMENT
# ============================================================================
def load_sheet_templates():
    """Load sheet templates from file, or use defaults with any saved custom ones merged.

    Only templates that were actually edited & saved (present as keys in
    sheet_templates.json) override the code default — every other template
    keeps running its default rule untouched. Deep-copied so no template's
    in-memory config can ever share a list/dict with DEFAULT_SHEETS and
    accidentally leak a later edit into the default.
    """
    if os.path.exists(TEMPLATES_FILE):
        try:
            with open(TEMPLATES_FILE, 'r') as f:
                saved_templates = json.load(f)
                merged = copy.deepcopy(DEFAULT_SHEETS)
                merged.update(saved_templates)  # only overrides templates present in the saved file
                return merged
        except:
            return copy.deepcopy(DEFAULT_SHEETS)
    return copy.deepcopy(DEFAULT_SHEETS)

def save_sheet_templates(templates):
    try:
        with open(TEMPLATES_FILE, 'w') as f:
            json.dump(templates, f, indent=2)
        return True
    except Exception as e:
        st.error(f"Error saving templates: {str(e)}")
        return False

def describe_template_rule(config):
    """Turn a template's raw config into a plain-English list of rules."""
    lines = []

    if not config.get('enabled', True):
        lines.append("⚠️ Disabled — this sheet will be skipped when generating reports.")

    ftype = config.get('filter_type', 'none')
    fval = config.get('filter_value', '')
    fvals = fval if isinstance(fval, list) else ([fval] if fval else [])

    if ftype == 'none':
        lines.append("✅ Starts with **all rows** (no primary include-filter).")
    elif ftype == 'remarks':
        lines.append(f"✅ Includes only rows where **Remarks** contains \"{fval}\".")
    elif ftype == 'sales_person_single':
        lines.append(f"✅ Includes only rows where **Sales Person Name** is exactly \"{fval}\".")
    elif ftype == 'sales_person_grouped':
        lines.append(f"✅ Includes only rows where **Sales Person Name** is exactly one of: {', '.join(fvals) if fvals else '(none set)'}.")
    elif ftype == 'line_grouped':
        lines.append(f"✅ Includes only rows where **Line** contains any of: {', '.join(fvals) if fvals else '(none set)'}. (Partial match — e.g. \"PHILIPS\" also matches \"PHILIPS-AIR FRYER\".)")
    else:
        lines.append(f"Filter type **{ftype}**, value: {fval}")

    for rule in (config.get('pre_processing_filters') or []):
        if not rule.get('enabled', True):
            continue
        col = rule.get('column', '?')
        txt = rule.get('contains', '')
        lines.append(f"❌ Excludes rows where **{col}** contains \"{txt}\".")

    excl = config.get('exclude_remarks')
    if excl:
        excl_vals = excl if isinstance(excl, list) else [excl]
        lines.append(f"❌ Excludes rows where **Remarks** contains any of: {', '.join(excl_vals)}.")

    if config.get('has_subtotals'):
        gb = config.get('group_by', 'Line')
        lines.append(f"📊 Grouped by **{gb}**, with a subtotal after each group and a grand TOTAL row at the end.")
    else:
        lines.append("📊 No grouping — plain row list.")

    lines.append(f"📋 Output columns: **{config.get('columns', 13)}-column** layout.")

    return lines

def load_custom_templates():
    if os.path.exists(CUSTOM_TEMPLATES_FILE):
        try:
            with open(CUSTOM_TEMPLATES_FILE, 'r') as f:
                return json.load(f)
        except:
            return {}
    return {}

def save_custom_templates(templates):
    try:
        with open(CUSTOM_TEMPLATES_FILE, 'w') as f:
            json.dump(templates, f, indent=2)
        return True
    except Exception as e:
        st.error(f"Error saving custom templates: {str(e)}")
        return False

if 'sheet_templates' not in st.session_state:
    st.session_state['sheet_templates'] = load_sheet_templates()

if 'custom_templates' not in st.session_state:
    st.session_state['custom_templates'] = load_custom_templates()

# ============================================================================
# MASTER FILE CREATION - WITH STOCK ALLOCATION LOGIC (PRODUCTION)
# ============================================================================
def convert_to_master_file(source_df, filters=None):
    """
    Convert raw file to Master file with STOCK ALLOCATION LOGIC.

    Production Rules:
    1. Apply filters to REMOVE unwanted rows BEFORE processing
    2. Group by Item Code (preferred) or Item Name
    3. Sort chronologically by date (oldest orders first)
    4. For each item:
       - Get original stock from first row
       - For each order (in date order):
         * Use PENDING QTY (not Order Qty!) for allocation
         * Calculate: Allocated Stock = min(remaining_stock, pending_qty)
         * Create tracking columns: Original, Allocated, After Order
         * Deduct from remaining_stock for next order

    Args:
        source_df: Raw DataFrame
        filters: List of filter rules [{'column': str, 'contains': str, 'enabled': bool}]

    Returns:
        DataFrame with stock allocation columns
    """
    df = source_df.copy()

    # ========================================================================
    # PRE-STEP: DETERMINE GROUPING COLUMN (needed for STEP 0)
    # ========================================================================
    group_col = None
    if 'Item Code' in source_df.columns:
        group_col = 'Item Code'
    elif 'Item Name' in source_df.columns:
        group_col = 'Item Name'
    else:
        # Can't proceed without grouping column
        st.warning("Neither 'Item Code' nor 'Item Name' found. Using raw data as master file.")
        return source_df

    # ========================================================================
    # STEP 0: READ ORIGINAL STOCK FROM SOURCE (BEFORE FILTERING)
    # ========================================================================
    # Critical: Store original stock from source data before any filtering
    # This ensures we have the correct stock value even if first row gets filtered
    original_stock_map = {}

    for item_key in source_df.groupby(group_col, sort=False).groups.keys():
        item_group = source_df[source_df[group_col].astype(str).str.strip() == str(item_key).strip()]

        if 'Stock Quantity' in item_group.columns:
            stock_values = item_group['Stock Quantity'].dropna()
            if len(stock_values) > 0:
                original_stock_map[item_key] = float(stock_values.iloc[0])

    # ========================================================================
    # STEP 1: APPLY PRE-PROCESSING FILTERS (REMOVE unwanted rows)
    # ========================================================================
    if filters:
        initial_count = len(df)
        total_removed = 0

        for filter_rule in filters:
            if not filter_rule.get('enabled', True):
                continue

            column = filter_rule.get('column')
            search_text = filter_rule.get('contains', '').strip()

            if not search_text or column not in df.columns:
                continue

            # Filter out (remove) matching rows
            before_count = len(df)
            mask = df[column].astype(str).str.lower().str.contains(search_text.lower(), na=False)
            df = df[~mask]  # Keep ONLY non-matching rows
            removed = before_count - len(df)

            if removed > 0:
                st.info(f"🔍 Filtered out {removed} row(s) where '{column}' contains '{search_text}'")
                total_removed += removed

        if total_removed > 0:
            final_count = len(df)
            st.success(f"✅ Total rows removed: {total_removed} | Remaining: {final_count}")

    # ========================================================================
    # STEP 2: DETECT DATE COLUMN
    # ========================================================================
    date_col = None
    date_col_display = None  # Store original date format column name

    for col in ['SO Date', 'Order Date', 'Date']:
        if col in df.columns:
            date_col = col
            date_col_display = f"{col}_DISPLAY"  # Create a display column
            break

    if date_col:
        # ✅ FIX: Keep dates as datetime objects for CORRECT SORTING
        # Convert to datetime (auto-detect format)
        try:
            df[date_col] = pd.to_datetime(df[date_col], format='%d-%m-%Y', errors='coerce', dayfirst=True)
        except:
            try:
                df[date_col] = pd.to_datetime(df[date_col], dayfirst=True, errors='coerce')
            except:
                pass  # Keep original if conversion fails

        # Create display column with DD-MM-YYYY format (for output only)
        if pd.api.types.is_datetime64_any_dtype(df[date_col]):
            df[date_col_display] = df[date_col].dt.strftime('%d-%m-%Y')

    # ========================================================================
    # STEP 3: SORT BY ITEM AND DATE (Chronological order)
    # ========================================================================
    # group_col already determined in PRE-STEP
    sort_cols = [group_col]
    if date_col:
        sort_cols.append(date_col)
    df = df.sort_values(sort_cols, ascending=[True, True])

    # ========================================================================
    # STEP 4: CALCULATE RUNNING STOCK ALLOCATION (FIFO METHOD WITH LAST ORDER RULE)
    # ========================================================================
    result_rows = []

    for item_key, group in df.groupby(group_col, sort=False):
        # ✅ FIX: EXPLICITLY SORT GROUP BY DATE ASCENDING (OLDEST FIRST) FOR FIFO
        # This ensures oldest orders are allocated stock first
        if date_col and date_col in group.columns:
            group = group.sort_values(by=date_col, ascending=True)

        # ✅ FIX: Use original stock from SOURCE data (before filtering)
        # This ensures correct allocation even if first row was filtered out
        if item_key in original_stock_map:
            original_stock = original_stock_map[item_key]
        else:
            # Fallback: read from group if not in map
            original_stock = 0
            if 'Stock Quantity' in group.columns:
                stock_values = group['Stock Quantity'].dropna()
                if len(stock_values) > 0:
                    original_stock = float(stock_values.iloc[0])

        remaining_stock = original_stock

        # Convert group to list to identify last order
        group_list = list(group.iterrows())
        total_orders = len(group_list)

        # ✅ FIFO: Process each order for this item (oldest first)
        for order_index, (idx, row) in enumerate(group_list):
            is_last_order = (order_index == total_orders - 1)

            # Get order quantity (PENDING QTY is priority - it's what still needs fulfilling)
            order_qty = 0
            if 'Pending Qty' in row.index:
                order_qty = float(row['Pending Qty']) if pd.notna(row['Pending Qty']) else 0
            elif 'Order Qty' in row.index:
                order_qty = float(row['Order Qty']) if pd.notna(row['Order Qty']) else 0

            # ✅ LAST ORDER RULE: Allocate ALL remaining stock
            if is_last_order:
                available_for_order = remaining_stock  # ⭐ ALL remaining stock
            else:
                # NON-LAST: Take only what's needed (FIFO)
                available_for_order = min(remaining_stock, order_qty) if remaining_stock > 0 else 0

            # Create row copy with tracking columns
            row_copy = row.copy()

            # Add tracking columns
            row_copy['Original Stock'] = original_stock
            row_copy['Allocated Stock'] = available_for_order
            row_copy['Stock After Order'] = max(0, remaining_stock - available_for_order)

            # Update Stock Quantity to show allocated stock
            if 'Stock Quantity' in row_copy.index:
                row_copy['Stock Quantity'] = available_for_order

            result_rows.append(row_copy)

            # Deduct from remaining stock for next order
            remaining_stock = max(0, remaining_stock - available_for_order)

    df_master = pd.DataFrame(result_rows)
    df_master = df_master.reset_index(drop=True)

    # ========================================================================
    # STEP 5: ADD PRODUCTION COLUMN (PRODUCTION = Pending Qty - Stock Quantity)
    # ========================================================================
    # PRODUCTION shows the production gap:
    # - PRODUCTION = 0: Fully covered by allocated stock (or last order with excess)
    # - PRODUCTION > 0: Need to produce X units to cover order
    if 'Pending Qty' in df_master.columns and 'Stock Quantity' in df_master.columns:
        df_master['PRODUCTION'] = df_master['Pending Qty'] - df_master['Stock Quantity']
        # ✅ Cap PRODUCTION at 0 (can't have negative production)
        # For last orders with excess stock, PRODUCTION = 0
        df_master['PRODUCTION'] = df_master['PRODUCTION'].apply(lambda x: max(0, x))
    else:
        df_master['PRODUCTION'] = 0

    # ========================================================================
    # STEP 6: SELECT ONLY REQUIRED 19 COLUMNS (MATCH TARGET FILE STRUCTURE)
    # ========================================================================
    # ✅ Fix: Replace datetime column with display format for output
    if date_col and date_col_display and date_col_display in df_master.columns:
        # Use the display format (DD-MM-YYYY) for output
        df_master[date_col] = df_master[date_col_display]
        df_master = df_master.drop(columns=[date_col_display])

    # Keep ONLY these 19 columns for final master file output
    # (Internal tracking columns are calculated but not exported)
    required_columns = [
        'SO NO.', 'SO Date', "Party's Order No.", 'Party Name', 'Item Name',
        'Item Code', 'Category', 'Base Unit', 'Order Qty', 'Dispatch Qty',
        'Pending Qty', 'Rate', 'Amount', 'Remarks', 'Stock Quantity',
        'PRODUCTION', 'Line', 'Delivery Date', 'Sales Person Name'
    ]

    # Keep only columns that exist in the dataframe
    available_cols = [col for col in required_columns if col in df_master.columns]
    df_master = df_master[available_cols]

    # ✅ Fill any remaining missing SO dates BEFORE returning master file
    df_master = fill_missing_so_dates(df_master)

    return df_master

def create_master_file(source_df, filters=None):
    """Wrapper for convert_to_master_file with error handling"""
    try:
        return convert_to_master_file(source_df, filters)
    except Exception as e:
        st.error(f"Error creating master file: {str(e)}")
        return None

# ============================================================================
# HELPER FUNCTIONS FOR SHEET GENERATION
# ============================================================================
def add_subtotal_row(df, total_label="Subtotal"):
    subtotal = {}
    for col in df.columns:
        if col in ['Order Qty', 'Dispatch Qty', 'Pending Qty', 'PRODUCTION']:
            try:
                subtotal[col] = df[col].sum()
            except:
                subtotal[col] = None
        else:
            subtotal[col] = total_label if col == 'Item Name' else ''
    return pd.DataFrame([subtotal])

def add_total_row(df, total_label="TOTAL"):
    total = {}
    for col in df.columns:
        if col in ['Order Qty', 'Pending Qty', 'PRODUCTION']:
            try:
                total[col] = df[col].sum()
            except:
                total[col] = None
        else:
            total[col] = total_label if col == 'Item Name' or col == 'Item Code' else ''
    return pd.DataFrame([total])

# ============================================================================
# ✅ FILL MISSING SO DATE VALUES - SENIOR DEVELOPER FIX
# ============================================================================
def fill_missing_so_dates(df):
    """
    Fill missing SO Date values intelligently:
    1. Identify subtotal rows (skip them - don't fill SO Date)
    2. Map SO NO to SO Date for actual data rows
    3. Use fallback to Order Date, Delivery Date, etc.
    4. Forward/backward fill for same SO NO
    5. Fill remaining nulls intelligently

    CRITICAL: Subtotal rows MUST remain with empty SO Date but are formatted blue
    Data rows MUST have SO Date populated
    """
    if 'SO Date' not in df.columns:
        return df

    df = df.copy()

    # IDENTIFY SUBTOTAL ROWS (mark them so we can skip them)
    subtotal_mask = df.apply(lambda row:
        any('Subtotal' in str(val) for val in row.values if pd.notna(val)),
        axis=1
    )

    # STEP 1: Map SO NO to its SO Date (only from DATA rows, not subtotals)
    if 'SO NO.' in df.columns:
        so_date_map = {}
        for so_no, group in df.groupby('SO NO.', sort=False):
            if pd.notna(so_no):  # Only map valid SO NOs
                # Exclude subtotal rows when creating map
                data_group = group[~subtotal_mask.loc[group.index]]
                valid_dates = data_group['SO Date'].dropna()
                if len(valid_dates) > 0:
                    so_date_map[str(so_no).strip()] = valid_dates.iloc[0]

        # STEP 2: Fill using SO NO map (but NOT subtotal rows)
        def fill_from_map(idx, row):
            # NEVER fill SO Date for subtotal rows
            if subtotal_mask.iloc[idx]:
                return row['SO Date']  # Keep empty

            if pd.isna(row['SO Date']) or str(row['SO Date']).strip() == '':
                so_no = str(row['SO NO.']).strip() if pd.notna(row['SO NO.']) else ''
                if so_no and so_no != 'nan':
                    return so_date_map.get(so_no, row['SO Date'])
            return row['SO Date']

        df['SO Date'] = df.apply(lambda row: fill_from_map(row.name, row), axis=1)

    # STEP 3: Use fallback columns for data rows only
    fallback_cols = ['Order Date', 'Delivery Date', "Party's Order Date", 'Date']
    for col in fallback_cols:
        if col in df.columns:
            # Only fill data rows, skip subtotals
            mask = (~subtotal_mask) & ((df['SO Date'].isna()) | (df['SO Date'] == ''))
            if mask.any():
                df.loc[mask, 'SO Date'] = df.loc[mask, col].fillna(df.loc[mask, 'SO Date'])

    # STEP 4: Forward/backward fill (data rows only)
    # Preserve original subtotal row positions
    data_indices = df.index[~subtotal_mask].tolist()
    df.loc[data_indices, 'SO Date'] = df.loc[data_indices, 'SO Date'].ffill().bfill()

    # STEP 5: Fill remaining blanks by looking backward (data rows only)
    for idx in df.index[~subtotal_mask]:
        if pd.isna(df.at[idx, 'SO Date']):
            # Look backward for previous non-null SO Date (from data rows)
            for prev_idx in range(idx-1, -1, -1):
                if not subtotal_mask.iloc[prev_idx] and pd.notna(df.at[prev_idx, 'SO Date']):
                    df.at[idx, 'SO Date'] = df.at[prev_idx, 'SO Date']
                    break

    # STEP 6: Final forward fill (data rows only)
    df.loc[data_indices, 'SO Date'] = df.loc[data_indices, 'SO Date'].ffill()

    # VERIFY: Subtotal rows must have empty SO Date
    df.loc[subtotal_mask, 'SO Date'] = ''

    return df

def sort_by_date(df):
    try:
        # Create a copy to avoid modifying original during sorting
        df_sort = df.copy()

        # ✅ Fill missing SO dates BEFORE sorting
        df_sort = fill_missing_so_dates(df_sort)

        if 'SO Date' in df_sort.columns:
            # Convert strings back to datetime for sorting only
            try:
                df_sort['_sort_key'] = pd.to_datetime(df_sort['SO Date'], format='%d-%m-%Y', errors='coerce', dayfirst=True)
                df_sort = df_sort.sort_values('_sort_key', ascending=False, na_position='last')
                df_sort = df_sort.drop('_sort_key', axis=1)
            except:
                df_sort = df_sort.sort_values('SO Date', ascending=False, na_position='last')
        elif 'Delivery Date' in df_sort.columns:
            # Convert strings back to datetime for sorting only
            try:
                df_sort['_sort_key'] = pd.to_datetime(df_sort['Delivery Date'], format='%d-%m-%Y', errors='coerce', dayfirst=True)
                df_sort = df_sort.sort_values('_sort_key', ascending=False, na_position='last')
                df_sort = df_sort.drop('_sort_key', axis=1)
            except:
                df_sort = df_sort.sort_values('Delivery Date', ascending=False, na_position='last')

        df_sort = df_sort.reset_index(drop=True)
        return df_sort
    except:
        return df

def sort_by_delivery_date_desc(df):
    """Sort by Delivery Date in descending order (newest first)"""
    try:
        df_sort = df.copy()

        # ✅ Fill missing SO dates
        df_sort = fill_missing_so_dates(df_sort)

        if 'Delivery Date' in df_sort.columns:
            # Convert strings back to datetime for sorting only
            try:
                df_sort['_sort_key'] = pd.to_datetime(df_sort['Delivery Date'], format='%d-%m-%Y', errors='coerce', dayfirst=True)
                df_sort = df_sort.sort_values('_sort_key', ascending=False, na_position='last')
                df_sort = df_sort.drop('_sort_key', axis=1)
            except:
                df_sort = df_sort.sort_values('Delivery Date', ascending=False, na_position='last')
        elif 'SO Date' in df_sort.columns:
            # Fallback to SO Date if Delivery Date not available
            try:
                df_sort['_sort_key'] = pd.to_datetime(df_sort['SO Date'], format='%d-%m-%Y', errors='coerce', dayfirst=True)
                df_sort = df_sort.sort_values('_sort_key', ascending=False, na_position='last')
                df_sort = df_sort.drop('_sort_key', axis=1)
            except:
                df_sort = df_sort.sort_values('SO Date', ascending=False, na_position='last')

        df_sort = df_sort.reset_index(drop=True)
        return df_sort
    except:
        return df

# ✅ Sort by Delivery Date ASCENDING (oldest first) within subtotal sections.
# Falls back to SO Date if Delivery Date is not available.
# NOTE: fill_missing_so_dates is intentionally NOT called here —
# calling it on a group with non-contiguous indices causes a silent IndexError.
def sort_by_so_date_asc(df):
    """Sort by Delivery Date ascending (oldest first). Pure sort — no date filling."""
    df_sort = df.copy()

    if 'Delivery Date' in df_sort.columns:
        df_sort['_sort_key'] = pd.to_datetime(
            df_sort['Delivery Date'], format='%d-%m-%Y', errors='coerce', dayfirst=True
        )
        df_sort = df_sort.sort_values('_sort_key', ascending=True, na_position='last')
        df_sort = df_sort.drop('_sort_key', axis=1)
    elif 'SO Date' in df_sort.columns:
        df_sort['_sort_key'] = pd.to_datetime(
            df_sort['SO Date'], format='%d-%m-%Y', errors='coerce', dayfirst=True
        )
        df_sort = df_sort.sort_values('_sort_key', ascending=True, na_position='last')
        df_sort = df_sort.drop('_sort_key', axis=1)

    return df_sort.reset_index(drop=True)

def apply_line_wise_subtotals_only(df):
    """
    Group by Line only.
    Add subtotals for each Line (no Sales Person level subtotals).
    Rows within each Line group sorted by SO Date ASCENDING.
    """
    if len(df) == 0:
        return df

    try:
        result_frames = []

        # ✅ FIX: Sort overall by SO Date ascending before grouping
        df = sort_by_so_date_asc(df)

        # Group by Line
        for line_name, line_group in df.groupby('Line', sort=False):
            # ✅ FIX: Sort each line group by SO Date ascending (oldest first)
            line_group = sort_by_so_date_asc(line_group)

            # Add all rows for this line
            result_frames.append(line_group)

            # Add Line subtotal only (no Sales Person level)
            if len(line_group) > 0:
                line_subtotal = add_subtotal_row(line_group, f"Subtotal: {line_name}")
                result_frames.append(line_subtotal)

            # Add separator
            separator = pd.DataFrame([['' for _ in df.columns]], columns=df.columns)
            result_frames.append(separator)

        result_df = pd.concat(result_frames, ignore_index=True)

        # Add GRAND TOTAL at the end
        if len(result_df) > 0:
            total_row = add_total_row(df, "GRAND TOTAL")
            result_df = pd.concat([result_df, total_row], ignore_index=True)

        return result_df
    except Exception as e:
        return df

def apply_column_set(sheet_df, num_cols):
    """Apply column set (13 or 19) - includes PRODUCTION in all sheets"""
    columns_13 = [
        'SO NO.', 'SO Date', "Party's Order No.", 'Item Name', 'Item Code',
        'Category', 'Base Unit', 'Order Qty', 'Remarks', 'PRODUCTION',
        'Line', 'Delivery Date', 'Sales Person Name'
    ]

    columns_19 = [
        'SO NO.', 'SO Date', "Party's Order No.", 'Party Name', 'Item Name',
        'Item Code', 'Category', 'Base Unit', 'Order Qty', 'Dispatch Qty',
        'Pending Qty', 'Rate', 'Amount', 'Remarks', 'Stock Quantity',
        'PRODUCTION', 'Line', 'Delivery Date', 'Sales Person Name'
    ]

    col_set = columns_13 if num_cols == 13 else columns_19
    available_cols = [c for c in col_set if c in sheet_df.columns]
    return sheet_df[available_cols]

# ============================================================================
# GENERATE SHEET WITH FILTER
# ============================================================================
def generate_sheet_with_filter(master_df, template_config):
    try:
        sheet_df = master_df.copy()

        # ✅ Fill missing SO dates FIRST - before any filtering
        sheet_df = fill_missing_so_dates(sheet_df)

        # ✅ Apply template pre-processing filters FIRST (Category, Remarks, etc.)
        if 'pre_processing_filters' in template_config and template_config['pre_processing_filters']:
            for filter_rule in template_config['pre_processing_filters']:
                column = filter_rule.get('column')
                contains_text = filter_rule.get('contains', '').strip()
                enabled = filter_rule.get('enabled', True)
                use_regex = filter_rule.get('regex', False)

                if enabled and column and contains_text and column in sheet_df.columns:
                    # Remove rows WHERE column CONTAINS the text (or matches regex if enabled)
                    if use_regex:
                        mask = sheet_df[column].astype(str).str.contains(contains_text, case=False, na=False, regex=True)
                    else:
                        mask = sheet_df[column].astype(str).str.contains(contains_text, case=False, na=False)
                    sheet_df = sheet_df[~mask]  # Keep rows that DON'T match

        # Apply filter
        if template_config['filter_type'] == 'remarks':
            sheet_df = sheet_df[sheet_df['Remarks'].str.contains(
                template_config['filter_value'], case=False, na=False)]

        elif template_config['filter_type'] == 'sales_person_single':
            # FIXED: Strip whitespace and case-insensitive matching
            filter_val = template_config['filter_value'].strip().upper()
            sheet_df = sheet_df[
                sheet_df['Sales Person Name'].astype(str).str.strip().str.upper() == filter_val
            ]

        elif template_config['filter_type'] == 'sales_person_grouped':
            sheet_df = sheet_df[sheet_df['Sales Person Name'].isin(template_config['filter_value'])]

            if template_config.get('has_subtotals') and len(sheet_df) > 0:
                # ✅ FIX: Sort by SO Date ascending within each Sales Person group
                sheet_df = sort_by_so_date_asc(sheet_df)
                grouped = sheet_df.groupby('Sales Person Name', sort=False)
                result_frames = []

                for person, group_df in grouped:
                    # ✅ FIX: Sort each group by SO Date ascending (oldest first)
                    group_df = sort_by_so_date_asc(group_df)
                    result_frames.append(group_df)
                    if len(group_df) > 0:
                        subtotal = add_subtotal_row(group_df, f"Subtotal: {person}")
                        result_frames.append(subtotal)

                sheet_df = pd.concat(result_frames, ignore_index=True)
                total_row = add_total_row(sheet_df, "TOTAL")
                sheet_df = pd.concat([sheet_df, total_row], ignore_index=True)
                return sheet_df

        elif template_config['filter_type'] == 'line_grouped':
            # FIXED: Handle both single string and list of filter values
            filter_vals = template_config['filter_value']
            if isinstance(filter_vals, str):
                filter_vals = [filter_vals]

            # Create mask for any matching value
            mask = pd.Series([False] * len(sheet_df), index=sheet_df.index)
            for fval in filter_vals:
                mask |= sheet_df['Line'].astype(str).str.contains(fval, case=False, na=False)
            sheet_df = sheet_df[mask]

            if template_config.get('has_subtotals') and len(sheet_df) > 0:
                # ✅ FIX: Sort by SO Date ascending within each Line group
                sheet_df = sort_by_so_date_asc(sheet_df)
                grouped = sheet_df.groupby('Line', sort=False)
                result_frames = []

                for line_name, group_df in grouped:
                    # ✅ FIX: Sort each group by SO Date ascending (oldest first)
                    group_df = sort_by_so_date_asc(group_df)
                    result_frames.append(group_df)
                    if len(group_df) > 0:
                        subtotal = add_subtotal_row(group_df, f"Subtotal: {line_name}")
                        result_frames.append(subtotal)

                sheet_df = pd.concat(result_frames, ignore_index=True)
                total_row = add_total_row(sheet_df, "TOTAL")
                sheet_df = pd.concat([sheet_df, total_row], ignore_index=True)
                return sheet_df

        # Apply exclude_remarks filter if specified
        if template_config.get('exclude_remarks'):
            exclude_vals = template_config['exclude_remarks']
            if isinstance(exclude_vals, str):
                exclude_vals = [exclude_vals]

            # Exclude rows where remarks contains any of the excluded values
            for exclude_val in exclude_vals:
                sheet_df = sheet_df[~sheet_df['Remarks'].astype(str).str.contains(
                    exclude_val, case=False, na=False)]

        # Apply Line-wise sorting and subtotals for all sheets
        if len(sheet_df) > 0:
            if 'Line' in sheet_df.columns:
                # Use line-wise grouping with line-only subtotals (SO Date ascending inside each group)
                sheet_df = apply_line_wise_subtotals_only(sheet_df)
            else:
                # Fallback: Just sort by SO Date ascending and add total
                sheet_df = sort_by_so_date_asc(sheet_df)
                total_row = add_total_row(sheet_df, "GRAND TOTAL")
                sheet_df = pd.concat([sheet_df, total_row], ignore_index=True)

        sheet_df = sheet_df.reset_index(drop=True)
        return sheet_df

    except Exception as e:
        st.error(f"Error generating sheet: {str(e)}")
        return None

# ============================================================================
# FORMAT DATE COLUMNS FOR EXCEL OUTPUT
# ============================================================================
def format_date_columns(df):
    """Convert all date columns to DD-MM-YYYY string format"""
    df = df.copy()

    date_columns = ['SO Date', 'Delivery Date', 'Order Date', 'Date']

    for col in date_columns:
        if col in df.columns:
            try:
                # Convert to datetime if not already
                df[col] = pd.to_datetime(df[col], format='%d-%m-%Y', errors='coerce', dayfirst=True)
                # Format as DD-MM-YYYY string
                df[col] = df[col].dt.strftime('%d-%m-%Y')
            except:
                pass

    return df

# ============================================================================
# ✅ FIX: DETECT CURRENT DATE ROWS (today, not yesterday)
# ============================================================================
def get_current_date():
    """Get today's date in DD-MM-YYYY format"""
    return datetime.now().strftime('%d-%m-%Y')

def is_row_current_date(row_data, sheet_df, row_index):
    """Check if a row's SO Date matches today"""
    today = get_current_date()

    # Check SO Date column only
    if 'SO Date' in sheet_df.columns:
        col_index = list(sheet_df.columns).index('SO Date')
        if col_index < len(row_data):
            cell_value = str(row_data[col_index]) if row_data[col_index] is not None else ""
            if cell_value == today:
                return True

    return False

# ============================================================================
# CREATE WORKBOOK WITH SHEETS
# ============================================================================
def create_workbook(master_df, sheet_configs, sheet_names_to_include=None):
    try:
        wb = Workbook()
        wb.remove(wb.active)

        report = {
            'total_sheets': 0,
            'generated_sheets': [],
            'issues': []
        }

        for sheet_name, config in sheet_configs.items():
            if sheet_names_to_include and sheet_name not in sheet_names_to_include:
                continue

            if not config.get('enabled', True):
                continue

            try:
                # Generate sheet data - Always use generate_sheet_with_filter to apply pre-processing filters
                sheet_df = generate_sheet_with_filter(master_df, config)

                if sheet_df is None or len(sheet_df) == 0:
                    report['issues'].append(f"{sheet_name}: No data after filtering")
                    continue

                # Apply column set
                sheet_df = apply_column_set(sheet_df, config.get('columns', 13))

                # Format date columns to DD-MM-YYYY BEFORE writing to Excel
                sheet_df = format_date_columns(sheet_df)

                # Create worksheet
                ws = wb.create_sheet(title=sheet_name[:31])

                # Header styling
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                # Add headers
                for col_num, col_name in enumerate(sheet_df.columns, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.value = col_name
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment

                # Pre-define fills and fonts
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                yellow_font = Font(color="000000", size=10)
                light_blue_fill = PatternFill(start_color="B4D7E8", end_color="B4D7E8", fill_type="solid")
                light_blue_font = Font(bold=True, size=10, color="000000")
                gold_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                gold_font = Font(bold=True, size=11, color="000000")

                # Add data
                for row_num, row_data in enumerate(dataframe_to_rows(sheet_df, index=False, header=False), 2):
                    # ✅ FIX: Detect row type at ROW level (not cell level)
                    # so ALL cells in the row get the correct highlight
                    is_today = is_row_current_date(row_data, sheet_df, row_num - 2)
                    is_subtotal_row = any(
                        'Subtotal' in str(v) for v in row_data if v is not None and str(v).strip() != ''
                    )
                    is_grand_total_row = any(
                        str(v).strip() in ('GRAND TOTAL', 'TOTAL') for v in row_data if v is not None and str(v).strip() != ''
                    )

                    for col_num, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.value = value

                        # ✅ Apply row-level styling (entire row gets the colour)
                        if is_today:
                            # Yellow highlight for today's date rows
                            cell.fill = yellow_fill
                            cell.font = yellow_font
                        elif is_grand_total_row:
                            # Gold background for GRAND TOTAL / TOTAL rows
                            cell.fill = gold_fill
                            cell.font = gold_font
                        elif is_subtotal_row:
                            # ✅ Light Blue for entire subtotal row
                            cell.fill = light_blue_fill
                            cell.font = light_blue_font

                # Auto-fit columns
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column].width = adjusted_width

                report['generated_sheets'].append({
                    'name': sheet_name,
                    'rows': len(sheet_df),
                    'columns': len(sheet_df.columns)
                })
                report['total_sheets'] += 1

            except Exception as e:
                report['issues'].append(f"{sheet_name}: {str(e)}")

        return wb, report

    except Exception as e:
        st.error(f"Error creating workbook: {str(e)}")
        return None, None

# ============================================================================
# MAIN APP UI
# ============================================================================

# Display system controls FIRST
show_system_status()

# Create layout with title and button
col_title, col_status, col_pull = st.columns([6, 2, 2])

with col_title:
    st.title("📊 Excel Master & Template Exporter v4.8")

with col_status:
    st.markdown("<br>", unsafe_allow_html=True)
    # Display last date/time on right side
    v = st.session_state.version_info
    if 'last_modified' in v:
        try:
            last_mod = datetime.fromisoformat(v['last_modified'])
            st.metric("📅 Last Update", last_mod.strftime("%d/%m/%y\n%H:%M"))
        except:
            st.metric("📅 Last Update", "N/A")
    else:
        st.metric("📅 Last Update", "N/A")

with col_pull:
    st.markdown("<br>", unsafe_allow_html=True)
    # Git Pull + Restart button (MAIN BUTTON)
    if st.button("🔄 PULL &\nRESTART", use_container_width=True, key="top_pull_restart", type="primary"):
        try:
            st.info("⏳ Starting Git pull and restart sequence...")

            # Step 1: Sync with GitHub
            progress_bar = st.progress(0)
            status_text = st.empty()

            status_text.text("📥 Pulling from GitHub...")
            progress_bar.progress(33)

            fetch_result = subprocess.run(
                ['git', 'fetch', 'origin', GITHUB_CONFIG['branch']],
                capture_output=True,
                text=True,
                timeout=30
            )

            pull_result = subprocess.run(
                ['git', 'pull', 'origin', GITHUB_CONFIG['branch']],
                capture_output=True,
                text=True,
                timeout=30
            )

            progress_bar.progress(66)
            status_text.text("✅ Code updated! Restarting Streamlit...")

            # Update version info
            st.session_state.version_info['last_modified'] = datetime.now().isoformat()
            save_version_info(st.session_state.version_info)
            add_changelog_entry("git_pull_restart", "Git pull + Streamlit restart", "Full deployment")

            progress_bar.progress(100)
            st.success("✅ Deployment complete! Restarting...")
            st.balloons()

            time.sleep(2)
            st.rerun()
        except Exception as e:
            st.error(f"❌ Error: {str(e)}")

# ============================================================================
# Optional: Also keep the simple restart button
# ============================================================================
show_restart_only = st.sidebar.checkbox("Show Restart Only Button", value=False)
if show_restart_only:
    if st.sidebar.button("🔁 Restart Only", use_container_width=True):
        st.sidebar.info("Restarting Streamlit...")
        st.session_state.version_info['last_modified'] = datetime.now().isoformat()
        save_version_info(st.session_state.version_info)
        add_changelog_entry("restart", "Streamlit restarted", "Manual restart")
        time.sleep(2)
        st.rerun()

# Version badge (NEW FEATURE)
with st.expander(f"ℹ️ **Version {get_version_string()} | Build {st.session_state.version_info['build']}**", expanded=False):
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Version", f"{st.session_state.version_info['major']}.{st.session_state.version_info['minor']}.{st.session_state.version_info['patch']}")
    with col2:
        st.metric("Build", st.session_state.version_info['build'])
    with col3:
        modified = datetime.fromisoformat(st.session_state.version_info['last_modified'])
        st.metric("Last Modified", modified.strftime("%d %b %y"))
    with col4:
        st.metric("Branch", get_current_branch())

    st.divider()
    st.markdown("**Recent Changes:**")
    changelog = load_changelog()[:5]
    for entry in changelog:
        ts = datetime.fromisoformat(entry['timestamp']).strftime("%Y-%m-%d %H:%M")
        st.caption(f"🔹 {entry['description']}\n_{ts}_")

# Show GitHub Dashboard if status button clicked
if st.session_state.get('show_git_status', False) and GITHUB_CONFIG['enabled']:
    show_github_dashboard()
    st.divider()
    if st.button("← Back to Main", use_container_width=True):
        st.session_state['show_git_status'] = False
        st.rerun()
    st.stop()

st.markdown("**10 Sheets | Combined + Separate Files | Template Management**")

tab1, tab2, tab3, tab_admin = st.tabs(["📥 Generate Sheets", "⚙️ Sheet Templates", "📁 Custom Templates", "🛠️ Admin & Deploy"])

# ============================================================================
# TAB 1: GENERATE SHEETS
# ============================================================================
with tab1:
    st.subheader("Step 1: Upload Source File")

    col1, col2 = st.columns([2, 1])

    with col1:
        source_file = st.file_uploader(
            "Upload Excel/CSV file",
            type=["xlsx", "xls", "csv"],
            key="source_upload"
        )

    with col2:
        if st.button("🔄 Reset", type="secondary", use_container_width=True):
            st.session_state.pop('master_file', None)
            st.session_state.pop('generated_workbook', None)
            st.session_state.pop('generation_report', None)
            st.success("Reset!")
            st.rerun()

    if source_file:
        try:
            if source_file.name.endswith('.csv'):
                source_df = pd.read_csv(source_file)
            else:
                source_df = pd.read_excel(source_file, sheet_name=0)

            st.success(f"✅ Loaded: {source_df.shape[0]} rows × {source_df.shape[1]} cols")

            with st.expander("📋 Preview Source Data"):
                st.dataframe(source_df.head(10), use_container_width=True)

            # ====================================================================
            # PRE-PROCESSING FILTERS
            # ====================================================================
            st.divider()
            st.subheader("🔍 Pre-Processing Filters (Remove unwanted rows)")
            st.caption("Filters are applied BEFORE master file creation to exclude unwanted data")

            if 'filters' not in st.session_state:
                st.session_state['filters'] = []

            col_filter1, col_filter2 = st.columns([3, 1])

            with col_filter1:
                st.markdown("**Add Filter:**")
                f_col1, f_col2, f_col3, f_col4 = st.columns([2, 2, 2, 1])

                with f_col1:
                    filter_column = st.selectbox(
                        "Column",
                        options=source_df.columns,
                        key="filter_col"
                    )

                with f_col2:
                    filter_contains = st.text_input(
                        "Contains (text)",
                        key="filter_text",
                        placeholder="e.g., 'Cancelled'"
                    )

                with f_col3:
                    filter_enabled = st.checkbox("Enabled", value=True, key="filter_enabled")

                with f_col4:
                    if st.button("➕ Add", use_container_width=True):
                        if filter_contains:
                            st.session_state['filters'].append({
                                'column': filter_column,
                                'contains': filter_contains,
                                'enabled': filter_enabled
                            })
                            st.rerun()

            with col_filter2:
                if st.button("🗑️ Clear All", type="secondary", use_container_width=True):
                    st.session_state['filters'] = []
                    st.rerun()

            # Display current filters
            if st.session_state['filters']:
                st.markdown("**Current Filters:**")
                for i, f in enumerate(st.session_state['filters']):
                    col_f1, col_f2, col_f3 = st.columns([3, 1, 1])
                    with col_f1:
                        status = "✅ Active" if f['enabled'] else "❌ Inactive"
                        st.write(f"{status}: Remove rows where '{f['column']}' contains '{f['contains']}'")
                    with col_f2:
                        if st.button("Toggle", key=f"toggle_{i}", use_container_width=True):
                            st.session_state['filters'][i]['enabled'] = not st.session_state['filters'][i]['enabled']
                            st.rerun()
                    with col_f3:
                        if st.button("❌", key=f"del_{i}", use_container_width=True):
                            st.session_state['filters'].pop(i)
                            st.rerun()

            # ====================================================================
            # CREATE MASTER FILE
            # ====================================================================
            st.divider()
            st.subheader("Step 2: Create Master File with Stock Allocation")
            st.caption("📊 Implements: Item grouping → Chronological sorting → Running stock allocation → Tracking columns")

            if st.button("🔄 Create Master File", type="primary", use_container_width=True):
                with st.spinner("Creating master file with stock allocation..."):
                    # Pass filters to master file creation
                    filters_to_apply = st.session_state['filters'] if st.session_state['filters'] else None
                    master_df = create_master_file(source_df, filters_to_apply)

                    if master_df is not None:
                        st.session_state['master_file'] = master_df
                        st.success("✅ Master file created with stock allocation!")

                        col_stat1, col_stat2, col_stat3 = st.columns(3)
                        with col_stat1:
                            st.metric("Rows", len(master_df))
                        with col_stat2:
                            st.metric("Columns", len(master_df.columns))
                        with col_stat3:
                            tracking_cols = ['Original Stock', 'Allocated Stock', 'Stock After Order']
                            has_tracking = sum(1 for col in tracking_cols if col in master_df.columns)
                            st.metric("Tracking Cols", has_tracking)

            if 'master_file' in st.session_state:
                with st.expander("📋 Preview Master File"):
                    master_df = st.session_state['master_file']

                    # Show tracking columns info
                    tracking_cols = ['Original Stock', 'Allocated Stock', 'Stock After Order']
                    available_tracking = [col for col in tracking_cols if col in master_df.columns]

                    if available_tracking:
                        st.info(f"✅ Tracking columns included: {', '.join(available_tracking)}")

                    st.dataframe(master_df.head(15), use_container_width=True)

                # ====================================================================
                # DOWNLOAD MASTER FILE
                # ====================================================================
                st.divider()
                st.subheader("📥 Download Master File")
                st.caption("Excel file with stock allocation, tracking columns, and PRODUCTION calculation")

                # Create master file Excel
                master_df = st.session_state['master_file']
                wb_master = Workbook()
                ws_master = wb_master.active
                ws_master.title = "Master File"

                # Header styling
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                # Add headers
                for col_num, col_name in enumerate(master_df.columns, 1):
                    cell = ws_master.cell(row=1, column=col_num)
                    cell.value = col_name
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment

                # Add data
                for row_num, row_data in enumerate(dataframe_to_rows(master_df, index=False, header=False), 2):
                    for col_num, value in enumerate(row_data, 1):
                        cell = ws_master.cell(row=row_num, column=col_num)
                        cell.value = value

                # Auto-fit columns
                for col in ws_master.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_master.column_dimensions[column].width = adjusted_width

                # Download button
                master_bytes = io.BytesIO()
                wb_master.save(master_bytes)
                master_bytes.seek(0)

                col_dl1, col_dl2 = st.columns([2, 1])
                with col_dl1:
                    st.download_button(
                        label="⬇️ Download Master File (.xlsx)",
                        data=master_bytes.getvalue(),
                        file_name=f"Master_File_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True
                    )

                with col_dl2:
                    st.metric("Rows", len(master_df))

        except Exception as e:
            st.error(f"Error: {str(e)}")

    # Generate Sheets
    st.divider()
    st.subheader("Step 3: Generate & Download Files")

    if 'master_file' not in st.session_state:
        st.warning("⚠️ Create Master File first")
    else:
        # Select sheets
        st.info("Select sheets to include in generated files:")

        cols = st.columns(5)
        selected_sheets = []

        for idx, (sheet_name, config) in enumerate(st.session_state['sheet_templates'].items()):
            col = cols[idx % 5]
            with col:
                if config.get('enabled', True):
                    if st.checkbox(sheet_name, value=True, key=f"select_{sheet_name}"):
                        selected_sheets.append(sheet_name)

        # Generate buttons
        col_gen1, col_gen2, col_gen3 = st.columns(3)

        with col_gen1:
            if st.button("📁 Generate COMBINED File", type="primary", use_container_width=True):
                with st.spinner("Generating combined workbook..."):
                    master_df = st.session_state['master_file']
                    wb, report = create_workbook(master_df, st.session_state['sheet_templates'], selected_sheets)

                    if wb and report:
                        st.session_state['generated_workbook_combined'] = wb
                        st.session_state['generation_report'] = report
                        st.success(f"✅ Generated {len(report['generated_sheets'])} sheets in combined file!")

        with col_gen2:
            if st.button("📂 Generate SEPARATE Files", type="primary", use_container_width=True):
                with st.spinner("Generating separate workbooks..."):
                    master_df = st.session_state['master_file']
                    st.session_state['separate_workbooks'] = {}

                    for sheet_name in selected_sheets:
                        config = st.session_state['sheet_templates'].get(sheet_name, {})
                        if not config.get('enabled', True):
                            continue

                        wb, _ = create_workbook(master_df, {sheet_name: config})
                        if wb:
                            st.session_state['separate_workbooks'][sheet_name] = wb

                    st.success(f"✅ Generated {len(st.session_state['separate_workbooks'])} separate files!")

        with col_gen3:
            if st.button("🎯 Generate BOTH", type="primary", use_container_width=True):
                with st.spinner("Generating all files..."):
                    master_df = st.session_state['master_file']

                    # Combined
                    wb_combined, report = create_workbook(master_df, st.session_state['sheet_templates'], selected_sheets)
                    st.session_state['generated_workbook_combined'] = wb_combined
                    st.session_state['generation_report'] = report

                    # Separate
                    st.session_state['separate_workbooks'] = {}
                    for sheet_name in selected_sheets:
                        config = st.session_state['sheet_templates'].get(sheet_name, {})
                        if not config.get('enabled', True):
                            continue

                        wb, _ = create_workbook(master_df, {sheet_name: config})
                        if wb:
                            st.session_state['separate_workbooks'][sheet_name] = wb

                    st.success(f"✅ Generated COMBINED + {len(st.session_state['separate_workbooks'])} SEPARATE files!")

        # Download sections
        st.divider()
        st.subheader("Step 4: Download Files")

        # Combined file
        if 'generated_workbook_combined' in st.session_state:
            st.markdown("### 📁 Combined File (All Sheets)")
            excel_bytes = io.BytesIO()
            st.session_state['generated_workbook_combined'].save(excel_bytes)
            excel_bytes.seek(0)

            st.download_button(
                label="⬇️ Download Combined Workbook (.xlsx)",
                data=excel_bytes.getvalue(),
                file_name=f"Master_All_Sheets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )

        # Separate files
        if 'separate_workbooks' in st.session_state and st.session_state['separate_workbooks']:
            st.markdown("### 📂 Separate Files (Individual Sheets)")

            col1, col2 = st.columns([1, 2])

            with col1:
                if st.button("📦 Download ALL as ZIP", type="secondary", use_container_width=True):
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                        for sheet_name, wb in st.session_state['separate_workbooks'].items():
                            excel_bytes = io.BytesIO()
                            wb.save(excel_bytes)
                            excel_bytes.seek(0)
                            zip_file.writestr(f"{sheet_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                            excel_bytes.getvalue())

                    zip_buffer.seek(0)
                    st.download_button(
                        label="⬇️ Download ZIP (All Sheets)",
                        data=zip_buffer.getvalue(),
                        file_name=f"Separate_Sheets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                        mime="application/zip",
                        type="secondary",
                        use_container_width=True
                    )

            # Individual downloads
            with col2:
                st.markdown("**Individual Sheet Downloads:**")
                for sheet_name, wb in st.session_state['separate_workbooks'].items():
                    col_a, col_b = st.columns([3, 1])
                    with col_b:
                        excel_bytes = io.BytesIO()
                        wb.save(excel_bytes)
                        excel_bytes.seek(0)

                        st.download_button(
                            label=f"⬇️ {sheet_name}",
                            data=excel_bytes.getvalue(),
                            file_name=f"{sheet_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"download_{sheet_name}",
                            use_container_width=True
                        )

# ============================================================================
# TAB 2: SHEET TEMPLATES
# ============================================================================
with tab2:
    st.subheader("⚙️ Manage Sheet Templates")

    st.info("Edit the default sheet configurations. Changes will affect next generation.")

    col_edit, col_view = st.columns([1, 1])

    with col_edit:
        st.markdown("### Edit Sheet")

        # If an Edit button was clicked in the view column, apply it before
        # instantiating the `edit_select` widget to avoid Streamlit's
        # "cannot modify after widget instantiated" error.
        if 'pending_edit' in st.session_state:
            st.session_state['edit_select'] = st.session_state.pop('pending_edit')

        sheet_to_edit = st.selectbox(
            "Select Sheet to Edit",
            options=list(st.session_state['sheet_templates'].keys()),
            key="edit_select"
        )

        if sheet_to_edit:
            config = st.session_state['sheet_templates'][sheet_to_edit]

            col_e1, col_e2 = st.columns(2)

            with col_e1:
                new_filter_type = st.selectbox(
                    "Filter Type",
                    options=['remarks', 'sales_person_single', 'sales_person_grouped', 'line_grouped', 'none'],
                    index=['remarks', 'sales_person_single', 'sales_person_grouped', 'line_grouped', 'none'].index(config.get('filter_type', 'none')),
                    key=f"filter_type_{sheet_to_edit}"
                )

            with col_e2:
                _current_filter_value = config.get('filter_value', '')
                if isinstance(_current_filter_value, list):
                    _filter_value_display = ', '.join(str(v) for v in _current_filter_value)
                else:
                    _filter_value_display = str(_current_filter_value)
                new_filter_value = st.text_input(
                    "Filter Value (comma-separated for multiple)",
                    value=_filter_value_display,
                    key=f"filter_value_{sheet_to_edit}"
                )

            col_e3, col_e4, col_e5 = st.columns(3)

            with col_e3:
                new_columns = st.selectbox(
                    "Columns",
                    options=[13, 19],
                    index=0 if config.get('columns', 13) == 13 else 1,
                    key=f"columns_{sheet_to_edit}"
                )

            with col_e4:
                new_subtotals = st.checkbox(
                    "Has Subtotals",
                    value=config.get('has_subtotals', False),
                    key=f"subtotals_{sheet_to_edit}"
                )

            with col_e5:
                new_enabled = st.checkbox(
                    "Enabled",
                    value=config.get('enabled', True),
                    key=f"enabled_{sheet_to_edit}"
                )

            new_desc = st.text_input(
                "Description",
                value=config.get('description', ''),
                key=f"desc_{sheet_to_edit}"
            )

            # Live preview of exactly what will be saved, so a bad paste is
            # visible before it's written to disk.
            if ',' in str(new_filter_value):
                _preview_val = [v.strip().strip("[]'\"") for v in str(new_filter_value).split(',')]
            else:
                _preview_val = str(new_filter_value).strip().strip("[]'\"")

            _preview_config = {
                **config,
                'filter_type': new_filter_type,
                'filter_value': _preview_val,
                'columns': new_columns,
                'has_subtotals': new_subtotals,
                'enabled': new_enabled,
            }
            st.markdown("**📝 Rule preview — this is what will be saved:**")
            for _line in describe_template_rule(_preview_config):
                st.markdown(f"- {_line}")

            col_save, col_reset = st.columns(2)

            with col_save:
                if st.button("💾 Save Changes", type="primary", use_container_width=True):
                    filter_val = _preview_val

                    st.session_state['sheet_templates'][sheet_to_edit] = {
                        'filter_type': new_filter_type,
                        'filter_value': filter_val,
                        'columns': new_columns,
                        'has_subtotals': new_subtotals,
                        'enabled': new_enabled,
                        'description': new_desc,
                        'group_by': config.get('group_by', 'Line')
                    }

                    save_sheet_templates(st.session_state['sheet_templates'])
                    st.success(f"✅ Updated {sheet_to_edit}! Filter value saved as: {filter_val}")
                    st.rerun()

            with col_reset:
                if sheet_to_edit in DEFAULT_SHEETS:
                    if st.button("🔄 Reset to Default", use_container_width=True):
                        st.session_state['sheet_templates'][sheet_to_edit] = copy.deepcopy(DEFAULT_SHEETS[sheet_to_edit])
                        save_sheet_templates(st.session_state['sheet_templates'])
                        st.success(f"✅ {sheet_to_edit} reset to code default!")
                        st.rerun()

    with col_view:
        st.markdown("### Current Templates")

        for name, config in st.session_state['sheet_templates'].items():
            status = "✅" if config.get('enabled', True) else "❌"
            st.write(f"{status} **{name}** — {config.get('description', '')}")

            for _line in describe_template_rule(config):
                st.markdown(f"- {_line}")

            with st.expander("Raw technical config"):
                filter_value = config.get('filter_value', '')
                filter_value_display = ', '.join(str(v) for v in filter_value) if isinstance(filter_value, list) else str(filter_value)
                st.caption(f"Filter: {config.get('filter_type', 'none')} | Value: {filter_value_display} | Cols: {config.get('columns', 13)} | Group by: {config.get('group_by', 'Line')}")

            # Provide quick-edit button that jumps to the Edit Sheet selector
            if st.button(f"Edit {name}", key=f"edit_btn_{name}"):
                st.session_state['pending_edit'] = name
                st.rerun()

            st.divider()

# ============================================================================
# TAB 3: CUSTOM TEMPLATES
# ============================================================================
with tab3:
    st.subheader("📁 Custom Excel Templates")

    col_custom1, col_custom2 = st.columns([1, 1])

    with col_custom1:
        st.markdown("### Create Custom Template")

        if 'master_file' not in st.session_state:
            st.warning("Create master file first in Tab 1")
        else:
            master_df = st.session_state['master_file']

            c_name = st.text_input("Template Name")
            c_cols = st.multiselect("Select Columns", options=master_df.columns)

            if c_cols and st.button("💾 Save Custom Template", type="primary", use_container_width=True):
                if not c_name:
                    st.error("Name required!")
                else:
                    st.session_state['custom_templates'][c_name] = {"columns": c_cols}
                    save_custom_templates(st.session_state['custom_templates'])
                    st.success(f"✅ Saved '{c_name}'!")
                    st.rerun()

    with col_custom2:
        st.markdown("### Manage Custom Templates")

        if not st.session_state['custom_templates']:
            st.info("No custom templates yet")
        else:
            selected = st.selectbox("Template", options=list(st.session_state['custom_templates'].keys()))

            if selected:
                cols = st.session_state['custom_templates'][selected].get('columns', [])
                if 'master_file' in st.session_state:
                    data = st.session_state['master_file'][[c for c in cols if c in st.session_state['master_file'].columns]]

                    st.metric("Rows", len(data))

                    st.download_button(
                        label="⬇️ Export CSV",
                        data=data.to_csv(index=False).encode(),
                        file_name=f"{selected}.csv",
                        mime="text/csv",
                        use_container_width=True
                    )

                    if st.button("🗑️ Delete", use_container_width=True):
                        del st.session_state['custom_templates'][selected]
                        save_custom_templates(st.session_state['custom_templates'])
                        st.rerun()

# ============================================================================
# TAB 4: ADMIN & DEPLOY (NEW)
# ============================================================================
with tab_admin:
    st.subheader("🛠️ Admin & Deployment Panel")
    st.info("⚠️ Operations that affect the running application")

    col_deploy1, col_deploy2 = st.columns([1, 1])

    with col_deploy1:
        st.markdown("### 🚀 Deployment Operations")

        # Main: Git Pull + Restart
        if st.button("🔄 Git Pull + Streamlit Restart", use_container_width=True, type="primary", key="admin_pull_restart"):
            try:
                progress_bar = st.progress(0)
                status_text = st.empty()

                status_text.text("📥 Step 1: Pulling from GitHub...")
                progress_bar.progress(25)

                fetch_result = subprocess.run(
                    ['git', 'fetch', 'origin', GITHUB_CONFIG['branch']],
                    capture_output=True,
                    text=True,
                    timeout=30
                )

                progress_bar.progress(50)
                status_text.text("📥 Step 2: Pulling latest changes...")

                pull_result = subprocess.run(
                    ['git', 'pull', 'origin', GITHUB_CONFIG['branch']],
                    capture_output=True,
                    text=True,
                    timeout=30
                )

                progress_bar.progress(75)
                status_text.text("✅ Code updated! Restarting Streamlit...")

                # Update version
                st.session_state.version_info['last_modified'] = datetime.now().isoformat()
                save_version_info(st.session_state.version_info)
                add_changelog_entry("git_pull_restart", "Git pull + Streamlit restart")

                progress_bar.progress(100)
                st.success("✅ Deployment Complete!")
                st.balloons()

                time.sleep(2)
                st.rerun()
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")

        st.divider()

        # Git Pull Only
        if st.button("📥 Git Pull Only", use_container_width=True, key="admin_pull_only"):
            try:
                progress_bar = st.progress(0)
                status_text = st.empty()

                status_text.text("📥 Pulling from GitHub...")
                progress_bar.progress(50)

                result = subprocess.run(
                    ['git', 'pull', 'origin', GITHUB_CONFIG['branch']],
                    capture_output=True,
                    text=True,
                    timeout=30
                )

                progress_bar.progress(100)

                if result.returncode == 0:
                    st.session_state.version_info['last_modified'] = datetime.now().isoformat()
                    save_version_info(st.session_state.version_info)
                    add_changelog_entry("git_pull", "Git pull completed")
                    st.success("✅ Git pull successful! (App not restarted)")
                    st.balloons()
                else:
                    st.error(f"❌ Git pull failed")
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")

        # Restart Only
        if st.button("🔁 Restart Streamlit Only", use_container_width=True, key="admin_restart_only"):
            try:
                progress_bar = st.progress(0)
                status_text = st.empty()

                status_text.text("🔄 Preparing restart...")
                progress_bar.progress(50)

                st.session_state.version_info['last_modified'] = datetime.now().isoformat()
                save_version_info(st.session_state.version_info)
                add_changelog_entry("restart", "Streamlit restarted")

                status_text.text("🚀 Restarting...")
                progress_bar.progress(100)

                st.success("✅ Restarting Streamlit...")
                st.balloons()

                time.sleep(2)
                st.rerun()
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")

        st.divider()

        # Push to GitHub
        st.markdown("### 📤 Push Changes")
        push_msg = st.text_input("Commit message", value="Update from Excel Exporter")

        if st.button("📤 Push to GitHub", use_container_width=True, key="admin_push"):
            try:
                progress_bar = st.progress(0)
                status_text = st.empty()

                status_text.text("📝 Staging changes...")
                progress_bar.progress(25)

                subprocess.run(['git', 'add', '.'], capture_output=True, timeout=10)

                progress_bar.progress(50)
                status_text.text("💾 Creating commit...")

                subprocess.run(['git', 'commit', '-m', push_msg],
                             capture_output=True, timeout=10)

                progress_bar.progress(75)
                status_text.text("📤 Pushing to GitHub...")

                result = subprocess.run(
                    ['git', 'push', 'origin', GITHUB_CONFIG['branch']],
                    capture_output=True,
                    text=True,
                    timeout=30
                )

                progress_bar.progress(100)

                if result.returncode == 0:
                    st.session_state.version_info['last_modified'] = datetime.now().isoformat()
                    save_version_info(st.session_state.version_info)
                    add_changelog_entry("git_push", f"Pushed to GitHub: {push_msg}")
                    st.success("✅ Pushed to GitHub!")
                    st.balloons()
                else:
                    st.warning("⚠️ Push completed with warnings")
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")

    with col_deploy2:
        st.markdown("### 📊 Application Status")

        # Version Information
        st.markdown("**Version Information:**")
        v = st.session_state.version_info

        st.metric("Version", f"{v['major']}.{v['minor']}.{v['patch']}")
        st.metric("Build", v['build'])

        try:
            created = datetime.fromisoformat(v['created'])
            st.caption(f"Created: {created.strftime('%Y-%m-%d %H:%M:%S')}")
        except:
            st.caption("Created: N/A")

        try:
            modified = datetime.fromisoformat(v['last_modified'])
            st.caption(f"Modified: {modified.strftime('%Y-%m-%d %H:%M:%S')}")
        except:
            st.caption("Modified: N/A")

        st.divider()

        # Git Status
        st.markdown("**Git Status:**")
        try:
            branch = get_current_branch()
            st.caption(f"🔗 Branch: {branch}")
        except:
            st.caption("🔗 Branch: unknown")

        try:
            commit = get_current_commit()
            st.caption(f"📌 Commit: {commit}")
        except:
            st.caption("📌 Commit: unknown")

        changes = get_uncommitted_changes()
        if changes > 0:
            st.warning(f"⚠️ {changes} file(s) with uncommitted changes")
        else:
            st.success("✅ Working directory is clean")

        st.divider()

        # Changelog
        st.markdown("**Changelog (Last 10):**")
        changelog = load_changelog()[:10]
        if changelog:
            for entry in changelog:
                try:
                    ts = datetime.fromisoformat(entry['timestamp']).strftime("%H:%M:%S")
                    type_emoji = {
                        "git_pull": "📥",
                        "git_push": "📤",
                        "restart": "🔁",
                        "git_pull_restart": "🚀",
                        "template_update": "📋",
                        "custom_template_update": "✨",
                        "export": "📊"
                    }.get(entry.get('type', 'other'), "📌")

                    st.caption(f"{type_emoji} {entry['description']}\n_{entry['version']} @ {ts}_")
                except:
                    st.caption(entry.get('description', 'Unknown'))
        else:
            st.caption("No changes logged yet")

# Footer
st.divider()
st.markdown("v4.8 | Combined + Separate Files | Template Management | One-Click Generation")
